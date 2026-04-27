import sys
import csv
import json
import re
import hashlib
import math
from bisect import bisect_right
from datetime import datetime
from pathlib import Path

import webview
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from services.export_service import ExportService
from services.gas_service import GasService
from services.lookup_service import LookupService
from services.project_service import ProjectService
from services.solar_service import SolarService
from services.wsfu_service import WsfuService
from version import APP_VERSION as PY_APP_VERSION


class Api:
    _APP_VERSION = PY_APP_VERSION
    _PROJECT_SHEET = "_project_data"
    _PROJECT_SCHEMA_VERSION = 1
    _GAS_MATERIALS = {
        "SCHEDULE_40_METALLIC",
        "CSST",
        "COPPER",
        "PE",
        "GALVANIZED_STEEL",
        "WROUGHT_IRON",
    }
    _SOLAR_CLIMATE_GROUPS = ("1,3,5,6", "2,4,6-14", "15")
    _SOLAR_BUILDING_TYPES = (
        "Grocery",
        "High Rise Multifamily",
        "Office, Financial Institutions, Unleased Tenant Space",
        "Retail",
        "School",
        "Conditoned Warehouse (to include mezzanine Sqft)",
        "Auditorium, Convention Center, Hotel/Motel, Library, Medical Office Building/Clinic, Restaurant, Theater",
        "Unconditioned Warehouse",
        "N/A",
    )
    _WSFU_FLUSH_TYPES = {"FLUSH_TANK", "FLUSH_VALVE"}
    _WSFU_PIPE_TYPE_C_FACTORS = [
        {"key": "L_COPPER", "label": "L COPPER", "cFactor": 130.0},
        {"key": "CPVC_SDR11", "label": "CPVC SDR 11", "cFactor": 150.0},
        {"key": "CPVC_SCH40", "label": "CPVC SCHEDULE 40", "cFactor": 150.0},
        {"key": "CPVC_SCH80", "label": "CPVC SCHEDULE 80", "cFactor": 145.0},
        {"key": "PP_SCH80", "label": "PP SCHEDULE 80", "cFactor": 145.0},
    ]
    _WSFU_SIZE_MAX_GPM = [
        {"size": '3/4"', "maxGpm": 8.0, "diameterIn": 0.82},
        {"size": '1"', "maxGpm": 13.0, "diameterIn": 1.05},
        {"size": '1-1/4"', "maxGpm": 24.0, "diameterIn": 1.38},
        {"size": '1-1/2"', "maxGpm": 35.0, "diameterIn": 1.61},
        {"size": '2"', "maxGpm": 62.0, "diameterIn": 2.07},
        {"size": '2-1/2"', "maxGpm": 95.0, "diameterIn": 2.47},
        {"size": '3"', "maxGpm": 140.0, "diameterIn": 3.07},
        {"size": '4"', "maxGpm": 245.0, "diameterIn": 4.03},
    ]

    def __init__(self) -> None:
        self._project_service = ProjectService()
        self._lookup_service = LookupService()
        self._wsfu_service = WsfuService()
        self._solar_service = SolarService()
        self._gas_service = GasService()
        self._export_service = ExportService()
        self._zip_index = {}
        self._data_error = ""
        self._wsfu_tank_curve = []
        self._wsfu_valve_curve = []
        self._wsfu_error = ""
        self._solar_zip_to_cz = {}
        self._lookup_service.load_lookup_data(self)
        self._wsfu_service.load_wsfu_data(self)
        self._solar_service.load_solar_zip_data(self)

    def _normalize_project_payload(self, payload: dict) -> dict:
        normalized = self._project_service.normalize_project_payload(
            payload,
            legacy_normalizer=self._normalize_project_payload_legacy,
            schema_version=self._PROJECT_SCHEMA_VERSION,
        )
        normalized["app_version"] = self._APP_VERSION
        return normalized

    def lookup_zip_rainfall(self, zip_code: str) -> dict:
        return self._lookup_service.lookup_zip_rainfall(self, zip_code)

    def calculate_wsfu(self, payload) -> dict:
        return self._wsfu_service.calculate_wsfu(self, payload)

    def get_solar_catalog(self) -> dict:
        return self._solar_service.get_solar_catalog(self)

    def lookup_solar_climate(self, zip_code: str) -> dict:
        return self._solar_service.lookup_solar_climate(self, zip_code)

    def calculate_solar(self, payload: dict) -> dict:
        return self._solar_service.calculate_solar(self, payload)

    def upload_gas_table_pdfs(self) -> dict:
        return self._gas_service.upload_gas_table_pdfs(self)

    def export_per_area(self, snapshot: dict) -> dict:
        return self._export_service.export_per_area(self, snapshot)

    def export_section(self, snapshot: dict) -> dict:
        return self._export_service.export_section(self, snapshot)

    def get_app_meta(self) -> dict:
        return {
            "ok": True,
            "app_version": self._APP_VERSION,
            "project_schema_version": self._PROJECT_SCHEMA_VERSION,
        }

    @staticmethod
    def _default_solar_factor_matrix() -> dict:
        # Workbook exposes one active scenario; seed all combinations with that baseline.
        baseline = {"A": 0.44, "B": 0.93, "C": 0.23}
        matrix = {}
        for bld in Api._SOLAR_BUILDING_TYPES:
            matrix[bld] = {}
            for group in Api._SOLAR_CLIMATE_GROUPS:
                matrix[bld][group] = dict(baseline)
        return matrix

    @staticmethod
    def _template_path() -> Path:
        candidates = [
            base_dir() / "app" / "Per_Area_Results_Template.xlsx",
            base_dir() / "Per_Area_Results_Template.xlsx",
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return candidates[0]

    @staticmethod
    def _lookup_file_path(filename: str) -> Path:
        candidates = [
            base_dir() / "app" / "data" / filename,
            base_dir() / "app" / filename,
            base_dir() / filename,
        ]
        for candidate in candidates:
            if candidate.exists():
                return candidate
        return candidates[0]

    @staticmethod
    def _normalize_zip(zip_code: str) -> str:
        digits = "".join(ch for ch in str(zip_code).strip() if ch.isdigit())
        if not digits:
            return ""
        # Preserve leading zeros for 5-digit U.S. ZIP codes.
        if len(digits) <= 5:
            return digits.zfill(5)
        return digits[:5]

    @staticmethod
    def _to_float(raw: str):
        try:
            return float(str(raw).strip())
        except Exception:
            return None

    @staticmethod
    def _round_up_one_decimal(value: float) -> float:
        # Always round upward to 1 decimal place: 1.19 -> 1.2
        return float((int(value * 10 + 0.9999999999)) / 10.0)

    @staticmethod
    def _to_non_negative_number(raw, fallback: float = 0.0) -> float:
        try:
            value = float(raw)
            if value < 0:
                return fallback
            return value
        except Exception:
            return fallback

    @staticmethod
    def _active_window():
        if getattr(webview, "windows", None) and len(webview.windows) > 0:
            return webview.windows[0]
        return None

    @staticmethod
    def _dialog_path(dialog_result):
        if not dialog_result:
            return ""
        if isinstance(dialog_result, (list, tuple)):
            if len(dialog_result) <= 0:
                return ""
            return str(dialog_result[0])
        return str(dialog_result)

    @staticmethod
    def _extract_export_project(snapshot: dict) -> dict:
        if not isinstance(snapshot, dict):
            return {}
        project_ctx = snapshot.get("project", {})
        if not isinstance(project_ctx, dict):
            return {}
        active = project_ctx.get("activeProject")
        if isinstance(active, dict):
            return active
        return {}

    @staticmethod
    def _write_project_header(sheet, start_row: int, project_info: dict) -> int:
        project_name = str(project_info.get("projectName", "") or "").strip()
        project_number = str(project_info.get("projectNumber", "") or "").strip()
        client_name = str(project_info.get("clientName", "") or "").strip()
        project_location = str(project_info.get("projectLocation", "") or "").strip()
        engineer_name = str(project_info.get("engineerName", "") or "").strip()
        company_name = str(project_info.get("companyName", "") or "").strip()
        project_date = str(project_info.get("date", "") or "").strip()
        revision = str(project_info.get("revision", "") or "").strip()

        if not (project_name or project_number or client_name or project_location):
            sheet.cell(row=start_row, column=1, value="Project Context")
            sheet.cell(row=start_row, column=2, value="No active project")
            return start_row + 2

        entries = [
            ("Project Name", project_name),
            ("Project Number", project_number),
            ("Client", client_name),
            ("Location", project_location),
            ("Date", project_date),
            ("Revision", revision),
            ("Engineer", engineer_name),
            ("Company", company_name),
        ]
        row = start_row
        for key, value in entries:
            if not value:
                continue
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)
            row += 1
        return row + 1

    def _normalize_project_payload_legacy(self, payload: dict) -> dict:
        if not isinstance(payload, dict):
            raise ValueError("Invalid project format.")

        schema_version = payload.get("schema_version", self._PROJECT_SCHEMA_VERSION)
        try:
            schema_version = int(schema_version)
        except Exception:
            raise ValueError("Invalid project schema version.")
        if schema_version != self._PROJECT_SCHEMA_VERSION:
            raise ValueError(
                f"Unsupported project schema version: {schema_version}. "
                f"Expected {self._PROJECT_SCHEMA_VERSION}."
            )

        settings = payload.get("settings", {})
        if not isinstance(settings, dict):
            settings = {}

        code_basis = str(settings.get("codeBasis", "IPC") or "IPC")
        if code_basis not in ("IPC", "CPC"):
            code_basis = "IPC"

        drain_type = str(settings.get("drainType", "PRIMARY") or "PRIMARY")
        if drain_type not in ("PRIMARY", "PRIMARY_SECONDARY", "COMBINED"):
            drain_type = "PRIMARY"

        slope = str(settings.get("slope", "1/8 in/ft") or "1/8 in/ft")
        zip_code = self._normalize_zip(str(settings.get("zipCode", "") or ""))
        rainfall_rate = self._to_non_negative_number(settings.get("rainfallRate", 1.0), 1.0)

        areas = payload.get("areas", [])
        if not isinstance(areas, list):
            areas = []
        normalized_areas = []
        for area in areas:
            if not isinstance(area, dict):
                continue
            area_name = str(area.get("name", "") or "").strip()
            main_area = self._to_non_negative_number(area.get("mainArea", 0), 0.0)
            sub_areas_raw = area.get("subAreas", [])
            if not isinstance(sub_areas_raw, list):
                sub_areas_raw = []
            side_walls_raw = area.get("sideWalls", [])
            if not isinstance(side_walls_raw, list):
                side_walls_raw = []
            sub_areas = [self._to_non_negative_number(v, 0.0) for v in sub_areas_raw]
            side_walls = [self._to_non_negative_number(v, 0.0) for v in side_walls_raw]
            normalized_areas.append(
                {
                    "name": area_name,
                    "mainArea": main_area,
                    "subAreas": sub_areas,
                    "sideWalls": side_walls,
                }
            )

        if not normalized_areas:
            normalized_areas = [{"name": "Area A", "mainArea": 0.0, "subAreas": [], "sideWalls": []}]
        else:
            for idx, area in enumerate(normalized_areas):
                if not str(area.get("name", "") or "").strip():
                    area["name"] = f"Area {chr(65+idx)}" if idx < 26 else f"Area {idx+1}"

        active_module = str(payload.get("activeModule", "home") or "home")
        if active_module not in ("home", "storm", "condensate", "vent", "ventilation", "gas", "wsfu", "fixtureUnit", "solar", "duct", "ductStatic", "refrigerant"):
            active_module = "home"

        def _normalize_project_entry(raw_entry):
            raw_entry = raw_entry if isinstance(raw_entry, dict) else {}
            return {
                "projectName": str(raw_entry.get("projectName", "") or "").strip(),
                "projectNumber": str(raw_entry.get("projectNumber", "") or "").strip(),
                "clientName": str(raw_entry.get("clientName", "") or "").strip(),
                "projectLocation": str(raw_entry.get("projectLocation", "") or "").strip(),
                "engineerName": str(raw_entry.get("engineerName", "") or "").strip(),
                "companyName": str(raw_entry.get("companyName", "") or "").strip(),
                "date": str(raw_entry.get("date", "") or "").strip(),
                "revision": str(raw_entry.get("revision", "") or "").strip(),
                "notes": str(raw_entry.get("notes", "") or "").strip(),
            }

        project_raw = payload.get("project", payload.get("projectContext", {}))
        if not isinstance(project_raw, dict):
            project_raw = {}
        project_active_raw = project_raw.get("activeProject")
        project_saved_raw = project_raw.get("savedProjects", [])
        if not isinstance(project_saved_raw, list):
            project_saved_raw = []

        normalized_saved_projects = []
        for project_item in project_saved_raw:
            if not isinstance(project_item, dict):
                continue
            normalized_saved_projects.append(_normalize_project_entry(project_item))

        normalized_active_project = _normalize_project_entry(project_active_raw) if isinstance(project_active_raw, dict) else None

        condensate_raw = payload.get("condensate", {})
        if not isinstance(condensate_raw, dict):
            condensate_raw = {}
        condensate_selection_raw = condensate_raw.get("selection", {})
        if not isinstance(condensate_selection_raw, dict):
            condensate_selection_raw = {}
        condensate_result_raw = condensate_raw.get("result", {})
        if not isinstance(condensate_result_raw, dict):
            condensate_result_raw = {}
        condensate_rows_raw = condensate_selection_raw.get("equipmentRows", [])
        if not isinstance(condensate_rows_raw, list):
            condensate_rows_raw = []
        normalized_condensate_rows = []
        for idx, row in enumerate(condensate_rows_raw):
            if not isinstance(row, dict):
                continue
            normalized_condensate_rows.append(
                {
                    "id": str(row.get("id", f"cond_{idx+1}") or f"cond_{idx+1}"),
                    "equipment": str(row.get("equipment", "AHU") or "AHU"),
                    "tonsPerUnit": self._to_non_negative_number(row.get("tonsPerUnit", 0.0), 0.0),
                    "quantity": self._to_non_negative_number(row.get("quantity", 1.0), 1.0),
                    "rowTotalTons": self._to_non_negative_number(row.get("rowTotalTons", 0.0), 0.0),
                    "rowSize": str(row.get("rowSize", "") or ""),
                }
            )
        if not normalized_condensate_rows:
            normalized_condensate_rows = [
                {
                    "id": "cond_1",
                    "equipment": str(condensate_selection_raw.get("equipment", "AHU") or "AHU"),
                    "tonsPerUnit": self._to_non_negative_number(condensate_selection_raw.get("tonsPerUnit", 0.0), 0.0),
                    "quantity": self._to_non_negative_number(condensate_selection_raw.get("quantity", 1.0), 1.0),
                    "rowTotalTons": 0.0,
                    "rowSize": "",
                }
            ]
        condensate_result_rows_raw = condensate_result_raw.get("rows", [])
        if not isinstance(condensate_result_rows_raw, list):
            condensate_result_rows_raw = []
        normalized_condensate_result_rows = []
        for idx, row in enumerate(condensate_result_rows_raw):
            if not isinstance(row, dict):
                continue
            normalized_condensate_result_rows.append(
                {
                    "id": str(row.get("id", f"cond_{idx+1}") or f"cond_{idx+1}"),
                    "equipment": str(row.get("equipment", "AHU") or "AHU"),
                    "equipmentLabel": str(row.get("equipmentLabel", "") or ""),
                    "tonsPerUnit": self._to_non_negative_number(row.get("tonsPerUnit", 0.0), 0.0),
                    "quantity": self._to_non_negative_number(row.get("quantity", 1.0), 1.0),
                    "rowTotalTons": self._to_non_negative_number(row.get("rowTotalTons", 0.0), 0.0),
                    "rowSize": str(row.get("rowSize", "") or ""),
                }
            )
        condensate_zones_raw = condensate_selection_raw.get("zones", [])
        if not isinstance(condensate_zones_raw, list):
            condensate_zones_raw = []
        normalized_condensate_zones = []
        for zidx, zone in enumerate(condensate_zones_raw):
            if not isinstance(zone, dict):
                continue
            zone_rows_raw = zone.get("equipmentRows", [])
            if not isinstance(zone_rows_raw, list):
                zone_rows_raw = []
            zone_rows = []
            for ridx, row in enumerate(zone_rows_raw):
                if not isinstance(row, dict):
                    continue
                zone_rows.append(
                    {
                        "id": str(row.get("id", f"cond_{zidx+1}_{ridx+1}") or f"cond_{zidx+1}_{ridx+1}"),
                        "equipment": str(row.get("equipment", "AHU") or "AHU"),
                        "tonsPerUnit": self._to_non_negative_number(row.get("tonsPerUnit", 0.0), 0.0),
                        "quantity": self._to_non_negative_number(row.get("quantity", 1.0), 1.0),
                        "rowTotalTons": self._to_non_negative_number(row.get("rowTotalTons", 0.0), 0.0),
                    }
                )
            if not zone_rows:
                zone_rows = [
                    {
                        "id": f"cond_{zidx+1}_1",
                        "equipment": "AHU",
                        "tonsPerUnit": 0.0,
                        "quantity": 1.0,
                        "rowTotalTons": 0.0,
                    }
                ]
            normalized_condensate_zones.append(
                {
                    "id": str(zone.get("id", f"zone_{zidx+1}") or f"zone_{zidx+1}"),
                    "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"),
                    "equipmentRows": zone_rows,
                }
            )
        if not normalized_condensate_zones:
            normalized_condensate_zones = [
                {
                    "id": "zone_1",
                    "name": "Zone 1",
                    "equipmentRows": normalized_condensate_rows,
                }
            ]

        condensate_result_zones_raw = condensate_result_raw.get("zones", [])
        if not isinstance(condensate_result_zones_raw, list):
            condensate_result_zones_raw = []
        normalized_condensate_result_zones = []
        for zidx, zone in enumerate(condensate_result_zones_raw):
            if not isinstance(zone, dict):
                continue
            zone_rows_raw = zone.get("rows", [])
            if not isinstance(zone_rows_raw, list):
                zone_rows_raw = []
            zone_rows = []
            for ridx, row in enumerate(zone_rows_raw):
                if not isinstance(row, dict):
                    continue
                zone_rows.append(
                    {
                        "id": str(row.get("id", f"cond_{zidx+1}_{ridx+1}") or f"cond_{zidx+1}_{ridx+1}"),
                        "equipment": str(row.get("equipment", "AHU") or "AHU"),
                        "equipmentLabel": str(row.get("equipmentLabel", "") or ""),
                        "tonsPerUnit": self._to_non_negative_number(row.get("tonsPerUnit", 0.0), 0.0),
                        "quantity": self._to_non_negative_number(row.get("quantity", 1.0), 1.0),
                        "rowTotalTons": self._to_non_negative_number(row.get("rowTotalTons", 0.0), 0.0),
                        "rowSize": str(row.get("rowSize", "") or ""),
                    }
                )
            normalized_condensate_result_zones.append(
                {
                    "id": str(zone.get("id", f"zone_{zidx+1}") or f"zone_{zidx+1}"),
                    "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"),
                    "rows": zone_rows,
                    "zoneTotalTons": self._to_non_negative_number(zone.get("zoneTotalTons", 0.0), 0.0),
                    "recommendedSize": str(zone.get("recommendedSize", "") or ""),
                    "warning": str(zone.get("warning", "") or ""),
                }
            )

        vent_raw = payload.get("vent", {})
        if not isinstance(vent_raw, dict):
            vent_raw = {}
        vent_selection_raw = vent_raw.get("selection", {})
        if not isinstance(vent_selection_raw, dict):
            vent_selection_raw = {}
        vent_result_raw = vent_raw.get("result", {})
        if not isinstance(vent_result_raw, dict):
            vent_result_raw = {}
        vent_rows_raw = vent_selection_raw.get("fixtureRows", [])
        if not isinstance(vent_rows_raw, list):
            vent_rows_raw = []
        normalized_vent_rows = []
        for idx, row in enumerate(vent_rows_raw):
            if not isinstance(row, dict):
                continue
            normalized_vent_rows.append(
                {
                    "id": str(row.get("id", f"vent_{idx+1}") or f"vent_{idx+1}"),
                    "fixtureKey": str(row.get("fixtureKey", "") or ""),
                    "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                    "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                    "unitDfu": self._to_non_negative_number(row.get("unitDfu", 0.0), 0.0),
                    "rowTotalDfu": self._to_non_negative_number(row.get("rowTotalDfu", 0.0), 0.0),
                }
            )
        sanitary_payload = payload.get("sanitary_drainage", {})
        if not isinstance(sanitary_payload, dict):
            sanitary_payload = {}
        vent_zones_raw = vent_selection_raw.get("zones", [])
        if not isinstance(vent_zones_raw, list):
            vent_zones_raw = []
        if not vent_zones_raw:
            vent_zones_raw = vent_raw.get("zones", [])
            if not isinstance(vent_zones_raw, list):
                vent_zones_raw = []
        if not vent_zones_raw:
            vent_zones_raw = sanitary_payload.get("zones", [])
            if not isinstance(vent_zones_raw, list):
                vent_zones_raw = []
        normalized_vent_zones = []
        for zidx, zone in enumerate(vent_zones_raw):
            if not isinstance(zone, dict):
                continue
            rows_raw = zone.get("fixtures", zone.get("fixtureRows", []))
            if not isinstance(rows_raw, list):
                rows_raw = []
            z_rows = []
            for ridx, row in enumerate(rows_raw):
                if not isinstance(row, dict):
                    continue
                z_rows.append(
                    {
                        "id": str(row.get("id", f"vent_{zidx+1}_{ridx+1}") or f"vent_{zidx+1}_{ridx+1}"),
                        "fixtureKey": str(row.get("fixtureKey", "") or ""),
                        "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                        "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                        "unitDfu": self._to_non_negative_number(row.get("unitDfu", 0.0), 0.0),
                        "rowTotalDfu": self._to_non_negative_number(row.get("rowTotalDfu", 0.0), 0.0),
                    }
                )
            if not z_rows:
                z_rows = [
                    {
                        "id": f"vent_{zidx+1}_1",
                        "fixtureKey": "",
                        "fixtureLabel": "",
                        "quantity": 0.0,
                        "unitDfu": 0.0,
                        "rowTotalDfu": 0.0,
                    }
                ]
            normalized_vent_zones.append(
                {
                    "id": str(zone.get("id", f"sd_zone_{zidx+1}") or f"sd_zone_{zidx+1}"),
                    "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"),
                    "fixtures": z_rows,
                }
            )
        if not normalized_vent_zones:
            normalized_vent_zones = [{"id": "sd_zone_1", "name": "Zone 1", "fixtures": normalized_vent_rows or [{"id": "vent_1", "fixtureKey": "", "fixtureLabel": "", "quantity": 0.0, "unitDfu": 0.0, "rowTotalDfu": 0.0}]}]

        ventilation_raw = payload.get("ventilation", {})
        if not isinstance(ventilation_raw, dict):
            ventilation_raw = {}
        ventilation_selection_raw = ventilation_raw.get("selection", {})
        if not isinstance(ventilation_selection_raw, dict):
            ventilation_selection_raw = {}
        ventilation_result_raw = ventilation_raw.get("result", {})
        if not isinstance(ventilation_result_raw, dict):
            ventilation_result_raw = {}
        ventilation_zones_raw = ventilation_selection_raw.get("zones", [])
        if not isinstance(ventilation_zones_raw, list):
            ventilation_zones_raw = []
        normalized_ventilation_zones = []
        for zidx, zone in enumerate(ventilation_zones_raw):
            if not isinstance(zone, dict):
                continue
            occupancy_key = str(zone.get("occupancyKey", "") or "")
            occupancy_category = str(zone.get("occupancyCategory", zone.get("occupancy_category", "")) or "")
            normalized_ventilation_zones.append(
                {
                    "id": str(zone.get("id", f"ventilation_zone_{zidx+1}") or f"ventilation_zone_{zidx+1}"),
                    "zoneName": str(zone.get("zoneName", f"Zone {zidx+1}") or f"Zone {zidx+1}"),
                    "spaceName": str(zone.get("spaceName", "") or ""),
                    "occupancyKey": occupancy_key or occupancy_category or "OFFICE_SPACE",
                    "occupancyCategory": occupancy_category,
                    "area": self._to_non_negative_number(zone.get("area", 0.0), 0.0),
                    "populationMode": "MANUAL" if str(zone.get("populationMode", "DEFAULT")) == "MANUAL" else "DEFAULT",
                    "manualPopulation": self._to_non_negative_number(zone.get("manualPopulation", 0.0), 0.0),
                    "ezPreset": str(zone.get("ezPreset", "DEFAULT_DESIGN_BASIS_08") or "DEFAULT_DESIGN_BASIS_08"),
                    "ez": self._to_non_negative_number(zone.get("ez", 0.8), 0.8),
                    "providedSupply": self._to_non_negative_number(zone.get("providedSupply", 0.0), 0.0),
                    "providedSupplySpecified": (
                        self._as_bool(zone.get("providedSupplySpecified", False), False)
                        or self._to_non_negative_number(zone.get("providedSupply", 0.0), 0.0) > 0.0
                    ),
                    "notes": str(zone.get("notes", "") or ""),
                }
            )
        gas_raw = payload.get("gas", {})
        if not isinstance(gas_raw, dict):
            gas_raw = {}

        def _normalize_code_basis(raw):
            value = str(raw or "IPC")
            if value not in ("IPC", "CPC"):
                return "IPC"
            return value

        def _normalize_fixture_category(raw):
            value = str(raw or "PRIVATE").upper()
            if value not in ("PRIVATE", "PUBLIC", "ASSEMBLY"):
                return "PRIVATE"
            return value

        def _normalize_fuel(raw):
            value = str(raw or "NATURAL_GAS")
            if value not in ("NATURAL_GAS", "PROPANE"):
                return "NATURAL_GAS"
            return value

        def _normalize_material(raw):
            value = str(raw or "SCHEDULE_40_METALLIC")
            if value not in self._GAS_MATERIALS:
                value = "SCHEDULE_40_METALLIC"
            return value

        gas_tables_raw = gas_raw.get("tables", [])
        if not isinstance(gas_tables_raw, list):
            gas_tables_raw = []
        normalized_tables = []
        for table in gas_tables_raw:
            if not isinstance(table, dict):
                continue
            table_id = str(table.get("id", "") or "").strip()
            if not table_id:
                table_id = hashlib.sha1(
                    f"{table.get('tableName', '')}-{table.get('sourceFile', '')}".encode("utf-8")
                ).hexdigest()[:12]
            length_rows_raw = table.get("lengthRows", [])
            if not isinstance(length_rows_raw, list):
                length_rows_raw = []
            length_rows = [self._to_non_negative_number(v, 0.0) for v in length_rows_raw]
            length_rows = [v for v in length_rows if v > 0]

            matrix_raw = table.get("sizeCapacityMatrix", {})
            if not isinstance(matrix_raw, dict):
                matrix_raw = {}
            size_matrix = {}
            for size_key, caps in matrix_raw.items():
                if not isinstance(caps, list):
                    continue
                size_matrix[str(size_key)] = [self._to_non_negative_number(v, 0.0) for v in caps]

            normalized_tables.append(
                {
                    "id": table_id,
                    "gasType": _normalize_fuel(table.get("gasType", "NATURAL_GAS")),
                    "inletPressure": self._to_non_negative_number(table.get("inletPressure", 0.0), 0.0),
                    "pressureDrop": self._to_non_negative_number(table.get("pressureDrop", 0.0), 0.0),
                    "specificGravity": self._to_non_negative_number(table.get("specificGravity", 0.6), 0.6),
                    "material": _normalize_material(table.get("material", "SCHEDULE_40_METALLIC")),
                    "codeBasis": _normalize_code_basis(table.get("codeBasis", "IPC")),
                    "sourceFile": str(table.get("sourceFile", "") or ""),
                    "tableName": str(table.get("tableName", "Gas Table") or "Gas Table"),
                    "lengthRows": length_rows,
                    "sizeCapacityMatrix": size_matrix,
                    "rawPreview": str(table.get("rawPreview", "") or ""),
                }
            )

        gas_selection_raw = gas_raw.get("selection", {})
        if not isinstance(gas_selection_raw, dict):
            gas_selection_raw = {}
        gas_lines_raw = gas_raw.get("lines", [])
        if not isinstance(gas_lines_raw, list):
            gas_lines_raw = []

        gas_result_raw = gas_raw.get("result", {})
        if not isinstance(gas_result_raw, dict):
            gas_result_raw = {}

        wsfu_raw = payload.get("wsfu", {})
        if not isinstance(wsfu_raw, dict):
            wsfu_raw = {}
        wsfu_selection_raw = wsfu_raw.get("selection", {})
        if not isinstance(wsfu_selection_raw, dict):
            wsfu_selection_raw = {}
        wsfu_result_raw = wsfu_raw.get("result", {})
        if not isinstance(wsfu_result_raw, dict):
            wsfu_result_raw = {}
        wsfu_rows_raw = wsfu_selection_raw.get("fixtureRows", [])
        if not isinstance(wsfu_rows_raw, list):
            wsfu_rows_raw = []
        normalized_wsfu_rows = []
        for row in wsfu_rows_raw:
            if not isinstance(row, dict):
                continue
            normalized_wsfu_rows.append(
                {
                    "id": str(row.get("id", "") or ""),
                    "fixtureKey": str(row.get("fixtureKey", "") or ""),
                    "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                    "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                    "hotCold": self._as_bool(row.get("hotCold", False), False),
                    "unitWsfu": self._to_non_negative_number(row.get("unitWsfu", 0.0), 0.0),
                    "rowTotalWsfu": self._to_non_negative_number(row.get("rowTotalWsfu", 0.0), 0.0),
                }
            )
        wsfu_zones_raw = wsfu_selection_raw.get("zones", [])
        if not isinstance(wsfu_zones_raw, list):
            wsfu_zones_raw = []
        if not wsfu_zones_raw:
            wsfu_zones_raw = wsfu_raw.get("zones", [])
            if not isinstance(wsfu_zones_raw, list):
                wsfu_zones_raw = []
        normalized_wsfu_zones = []
        for zidx, zone in enumerate(wsfu_zones_raw):
            if not isinstance(zone, dict):
                continue
            rows_raw = zone.get("fixtures", zone.get("fixtureRows", []))
            if not isinstance(rows_raw, list):
                rows_raw = []
            z_rows = []
            for ridx, row in enumerate(rows_raw):
                if not isinstance(row, dict):
                    continue
                z_rows.append(
                    {
                        "id": str(row.get("id", f"wsfu_{zidx+1}_{ridx+1}") or f"wsfu_{zidx+1}_{ridx+1}"),
                        "fixtureKey": str(row.get("fixtureKey", "") or ""),
                        "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                        "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                        "hotCold": self._as_bool(row.get("hotCold", False), False),
                        "unitWsfu": self._to_non_negative_number(row.get("unitWsfu", 0.0), 0.0),
                        "rowTotalWsfu": self._to_non_negative_number(row.get("rowTotalWsfu", 0.0), 0.0),
                    }
                )
            if not z_rows:
                z_rows = [{"id": f"wsfu_{zidx+1}_1", "fixtureKey": "", "fixtureLabel": "", "quantity": 0.0, "hotCold": False, "unitWsfu": 0.0, "rowTotalWsfu": 0.0}]
            normalized_wsfu_zones.append({"id": str(zone.get("id", f"wsfu_zone_{zidx+1}") or f"wsfu_zone_{zidx+1}"), "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"), "fixtures": z_rows})
        if not normalized_wsfu_zones:
            normalized_wsfu_zones = [{"id": "wsfu_zone_1", "name": "Zone 1", "fixtures": normalized_wsfu_rows or [{"id": "wsfu_1", "fixtureKey": "", "fixtureLabel": "", "quantity": 0.0, "hotCold": False, "unitWsfu": 0.0, "rowTotalWsfu": 0.0}]}]

        fixture_unit_raw = payload.get("fixtureUnit", {})
        if not isinstance(fixture_unit_raw, dict):
            fixture_unit_raw = {}
        fixture_unit_selection_raw = fixture_unit_raw.get("selection", {})
        if not isinstance(fixture_unit_selection_raw, dict):
            fixture_unit_selection_raw = {}
        fixture_unit_result_raw = fixture_unit_raw.get("result", {})
        if not isinstance(fixture_unit_result_raw, dict):
            fixture_unit_result_raw = {}
        fixture_unit_rows_raw = fixture_unit_selection_raw.get("rows", [])
        if not isinstance(fixture_unit_rows_raw, list):
            fixture_unit_rows_raw = []
        fixture_unit_zones_raw = fixture_unit_selection_raw.get("zones", [])
        if not isinstance(fixture_unit_zones_raw, list):
            fixture_unit_zones_raw = []
        normalized_fixture_unit_rows = []
        for ridx, row in enumerate(fixture_unit_rows_raw):
            if not isinstance(row, dict):
                continue
            normalized_fixture_unit_rows.append(
                {
                    "id": str(row.get("id", f"fixture_unit_row_{ridx+1}") or f"fixture_unit_row_{ridx+1}"),
                    "fixtureKey": str(row.get("fixtureKey", "") or ""),
                    "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                    "category": _normalize_fixture_category(row.get("category", "PRIVATE")),
                    "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                    "hotCold": self._as_bool(row.get("hotCold", False), False),
                    "note": str(row.get("note", "") or ""),
                    "wasteUnit": self._to_non_negative_number(
                        row.get("wasteUnit", row.get("sanitaryUnit", 0.0)), 0.0
                    ),
                    "waterUnit": self._to_non_negative_number(row.get("waterUnit", 0.0), 0.0),
                    "rowWasteTotal": self._to_non_negative_number(
                        row.get("rowWasteTotal", row.get("rowSanitaryTotal", 0.0)), 0.0
                    ),
                    "rowWaterTotal": self._to_non_negative_number(row.get("rowWaterTotal", 0.0), 0.0),
                }
            )
        if not normalized_fixture_unit_rows:
            normalized_fixture_unit_rows = [
                {
                    "id": "fixture_unit_row_1",
                    "fixtureKey": "",
                    "fixtureLabel": "",
                    "category": "PRIVATE",
                    "quantity": 0.0,
                    "hotCold": False,
                    "note": "",
                    "wasteUnit": 0.0,
                    "waterUnit": 0.0,
                    "rowWasteTotal": 0.0,
                    "rowWaterTotal": 0.0,
                }
            ]
        normalized_fixture_unit_zones = []
        for zidx, zone in enumerate(fixture_unit_zones_raw):
            if not isinstance(zone, dict):
                continue
            rows_raw = zone.get("rows", [])
            if not isinstance(rows_raw, list):
                rows_raw = []
            z_rows = []
            for ridx, row in enumerate(rows_raw):
                if not isinstance(row, dict):
                    continue
                z_rows.append(
                    {
                        "id": str(row.get("id", f"fixture_unit_{zidx+1}_{ridx+1}") or f"fixture_unit_{zidx+1}_{ridx+1}"),
                        "fixtureKey": str(row.get("fixtureKey", "") or ""),
                        "fixtureLabel": str(row.get("fixtureLabel", "") or ""),
                        "category": _normalize_fixture_category(row.get("category", "PRIVATE")),
                        "quantity": self._to_non_negative_number(row.get("quantity", 0.0), 0.0),
                        "hotCold": self._as_bool(row.get("hotCold", False), False),
                        "note": str(row.get("note", "") or ""),
                        "wasteUnit": self._to_non_negative_number(
                            row.get("wasteUnit", row.get("sanitaryUnit", 0.0)), 0.0
                        ),
                        "waterUnit": self._to_non_negative_number(row.get("waterUnit", 0.0), 0.0),
                        "rowWasteTotal": self._to_non_negative_number(
                            row.get("rowWasteTotal", row.get("rowSanitaryTotal", 0.0)), 0.0
                        ),
                        "rowWaterTotal": self._to_non_negative_number(row.get("rowWaterTotal", 0.0), 0.0),
                    }
                )
            if not z_rows:
                z_rows = [
                    {
                        "id": f"fixture_unit_{zidx+1}_1",
                        "fixtureKey": "",
                        "fixtureLabel": "",
                        "category": "PRIVATE",
                        "quantity": 0.0,
                        "hotCold": False,
                        "note": "",
                        "wasteUnit": 0.0,
                        "waterUnit": 0.0,
                        "rowWasteTotal": 0.0,
                        "rowWaterTotal": 0.0,
                    }
                ]
            normalized_fixture_unit_zones.append(
                {
                    "id": str(zone.get("id", f"fixture_unit_zone_{zidx+1}") or f"fixture_unit_zone_{zidx+1}"),
                    "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"),
                    "rows": z_rows,
                }
            )
        if not normalized_fixture_unit_zones:
            normalized_fixture_unit_zones = [
                {"id": "fixture_unit_zone_1", "name": "Zone 1", "rows": normalized_fixture_unit_rows}
            ]

        duct_raw = payload.get("duct", {})
        if not isinstance(duct_raw, dict):
            duct_raw = {}
        duct_selection_raw = duct_raw.get("selection", {})
        if not isinstance(duct_selection_raw, dict):
            duct_selection_raw = {}
        duct_result_raw = duct_raw.get("result", {})
        if not isinstance(duct_result_raw, dict):
            duct_result_raw = {}
        duct_sizing_payload = payload.get("duct_sizing", {})
        if not isinstance(duct_sizing_payload, dict):
            duct_sizing_payload = {}
        duct_zones_raw = duct_selection_raw.get("zones", [])
        if not isinstance(duct_zones_raw, list):
            duct_zones_raw = []
        if not duct_zones_raw:
            duct_zones_raw = duct_raw.get("zones", [])
            if not isinstance(duct_zones_raw, list):
                duct_zones_raw = []
        if not duct_zones_raw:
            duct_zones_raw = duct_sizing_payload.get("zones", [])
            if not isinstance(duct_zones_raw, list):
                duct_zones_raw = []
        normalized_duct_zones = []
        for zidx, zone in enumerate(duct_zones_raw):
            if not isinstance(zone, dict):
                continue
            segments_raw = zone.get("segments", [])
            if not isinstance(segments_raw, list):
                segments_raw = []
            z_segments = []
            for sidx, seg in enumerate(segments_raw):
                if not isinstance(seg, dict):
                    continue
                z_segments.append(
                    {
                        "id": str(seg.get("id", f"duct_{zidx+1}_{sidx+1}") or f"duct_{zidx+1}_{sidx+1}"),
                        "name": str(seg.get("name", f"Segment {sidx+1}") or f"Segment {sidx+1}"),
                        "mode": str(seg.get("mode", "SIZE_FROM_CFM_FRICTION") or "SIZE_FROM_CFM_FRICTION"),
                        "shape": str(seg.get("shape", "ROUND") or "ROUND"),
                        "cfm": self._to_non_negative_number(seg.get("cfm", 0.0), 0.0),
                        "frictionTarget": self._to_non_negative_number(seg.get("frictionTarget", 0.0), 0.0),
                        "velocityTarget": self._to_non_negative_number(seg.get("velocityTarget", 0.0), 0.0),
                        "diameter": self._to_non_negative_number(seg.get("diameter", 0.0), 0.0),
                        "width": self._to_non_negative_number(seg.get("width", 0.0), 0.0),
                        "height": self._to_non_negative_number(seg.get("height", 0.0), 0.0),
                        "ratioLimit": self._to_non_negative_number(seg.get("ratioLimit", 3.0), 3.0),
                    }
                )
            if not z_segments:
                z_segments = [{"id": f"duct_{zidx+1}_1", "name": "Segment 1", "mode": "SIZE_FROM_CFM_FRICTION", "shape": "ROUND", "cfm": 0.0, "frictionTarget": 0.0, "velocityTarget": 0.0, "diameter": 0.0, "width": 0.0, "height": 0.0, "ratioLimit": 3.0}]
            normalized_duct_zones.append({"id": str(zone.get("id", f"duct_zone_{zidx+1}") or f"duct_zone_{zidx+1}"), "name": str(zone.get("name", f"Zone {zidx+1}") or f"Zone {zidx+1}"), "segments": z_segments})
        if not normalized_duct_zones:
            normalized_duct_zones = [{
                "id": "duct_zone_1",
                "name": "Zone 1",
                "segments": [{
                    "id": "duct_1_1",
                    "name": "Segment 1",
                    "mode": str(duct_selection_raw.get("mode", "SIZE_FROM_CFM_FRICTION") or "SIZE_FROM_CFM_FRICTION"),
                    "shape": str(duct_selection_raw.get("shape", "ROUND") or "ROUND"),
                    "cfm": self._to_non_negative_number(duct_selection_raw.get("cfm", 0.0), 0.0),
                    "frictionTarget": self._to_non_negative_number(duct_selection_raw.get("frictionTarget", 0.0), 0.0),
                    "velocityTarget": self._to_non_negative_number(duct_selection_raw.get("velocityTarget", 0.0), 0.0),
                    "diameter": self._to_non_negative_number(duct_selection_raw.get("diameter", 0.0), 0.0),
                    "width": self._to_non_negative_number(duct_selection_raw.get("width", 0.0), 0.0),
                    "height": self._to_non_negative_number(duct_selection_raw.get("height", 0.0), 0.0),
                    "ratioLimit": self._to_non_negative_number(duct_selection_raw.get("ratioLimit", 3.0), 3.0),
                }],
            }]

        duct_static_raw = payload.get("ductStatic", {})
        if not isinstance(duct_static_raw, dict):
            duct_static_raw = {}
        duct_static_selection_raw = duct_static_raw.get("selection", {})
        if not isinstance(duct_static_selection_raw, dict):
            duct_static_selection_raw = {}
        duct_static_result_raw = duct_static_raw.get("result", {})
        if not isinstance(duct_static_result_raw, dict):
            duct_static_result_raw = {}
        duct_static_segments_raw = duct_static_selection_raw.get("segments", [])
        if not isinstance(duct_static_segments_raw, list):
            duct_static_segments_raw = []
        normalized_duct_static_segments = []
        for seg in duct_static_segments_raw:
            if not isinstance(seg, dict):
                continue
            line_type = str(seg.get("lineType", "Straight") or "Straight")
            if line_type not in ("Straight", "Elbow", "Transition", "Outlet"):
                line_type = "Straight"
            normalized_duct_static_segments.append(
                {
                    "id": str(seg.get("id", "") or ""),
                    "name": str(seg.get("name", "") or ""),
                    "lineType": line_type,
                    "lengthFt": self._to_non_negative_number(seg.get("lengthFt", 0.0), 0.0),
                    "frictionRate": self._to_non_negative_number(seg.get("frictionRate", 0.0), 0.0),
                    "width": self._to_non_negative_number(seg.get("width", 0.0), 0.0),
                    "height": self._to_non_negative_number(seg.get("height", 0.0), 0.0),
                    "radius": self._to_non_negative_number(seg.get("radius", 0.0), 0.0),
                    "velocityFpm": self._to_non_negative_number(seg.get("velocityFpm", 0.0), 0.0),
                    "rwRatio": self._to_non_negative_number(seg.get("rwRatio", 0.0), 0.0),
                    "hwRatio": self._to_non_negative_number(seg.get("hwRatio", 0.0), 0.0),
                    "cCoefficient": self._to_non_negative_number(seg.get("cCoefficient", 0.0), 0.0),
                    "segmentPressureDrop": self._to_non_negative_number(seg.get("segmentPressureDrop", 0.0), 0.0),
                    "warning": str(seg.get("warning", "") or ""),
                }
            )

        solar_raw = payload.get("solar", {})
        if not isinstance(solar_raw, dict):
            solar_raw = {}
        solar_selection_raw = solar_raw.get("selection", {})
        if not isinstance(solar_selection_raw, dict):
            solar_selection_raw = {}
        solar_result_raw = solar_raw.get("result", {})
        if not isinstance(solar_result_raw, dict):
            solar_result_raw = {}

        refrigerant_raw = payload.get("refrigerant", {})
        if not isinstance(refrigerant_raw, dict):
            refrigerant_raw = {}
        refrigerant_selection_raw = refrigerant_raw.get("selection", {})
        if not isinstance(refrigerant_selection_raw, dict):
            refrigerant_selection_raw = {}
        refrigerant_result_raw = refrigerant_raw.get("result", {})
        if not isinstance(refrigerant_result_raw, dict):
            refrigerant_result_raw = {}

        # Legacy fallback support (older gas payload shape).
        if not gas_selection_raw:
            gas_selection_raw = {
                "codeBasis": gas_raw.get("codeBasis", "IPC"),
                "gasType": gas_raw.get("fuelType", "NATURAL_GAS"),
                "longestLength": gas_raw.get("longestLength", 40),
                "demandMode": "CFH",
                "demandValue": gas_raw.get("demandCfh", 100),
                "heatingValue": 1000 if str(gas_raw.get("fuelType", "NATURAL_GAS")) == "NATURAL_GAS" else 2500,
                "inletPressure": gas_raw.get("inletPressure", 0.0),
                "pressureDrop": gas_raw.get("pressureDrop", 0.0),
                "specificGravity": gas_raw.get("specificGravity", 0.6),
                "selectedTableId": gas_raw.get("selectedTableId", ""),
            }
            gas_lines_raw = [
                {
                    "id": "line_1",
                    "label": "Line 1",
                    "demandValue": gas_raw.get("demandCfh", 100),
                    "runLength": gas_raw.get("longestLength", 40),
                }
            ]

        if not gas_result_raw and isinstance(gas_raw.get("result", {}), dict):
            gas_result_raw = gas_raw.get("result", {})

        normalized_lines = []
        for idx, line in enumerate(gas_lines_raw):
            if not isinstance(line, dict):
                continue
            normalized_lines.append(
                {
                    "id": str(line.get("id", f"line_{idx+1}") or f"line_{idx+1}"),
                    "label": str(line.get("label", f"Line {idx+1}") or f"Line {idx+1}"),
                    "demandValue": self._to_non_negative_number(line.get("demandValue", 0), 0.0),
                    "runLength": self._to_non_negative_number(line.get("runLength", 0), 0.0),
                }
            )
        if not normalized_lines:
            normalized_lines = [{"id": "line_1", "label": "Line 1", "demandValue": 100.0, "runLength": 40.0}]

        result_lines_raw = gas_result_raw.get("lines", [])
        if not isinstance(result_lines_raw, list):
            result_lines_raw = []
        normalized_result_lines = []
        for line in result_lines_raw:
            if not isinstance(line, dict):
                continue
            normalized_result_lines.append(
                {
                    "lineId": str(line.get("lineId", "") or ""),
                    "lineLabel": str(line.get("lineLabel", "") or ""),
                    "demandCfh": self._to_non_negative_number(line.get("demandCfh", 0), 0.0),
                    "runLength": self._to_non_negative_number(line.get("runLength", 0), 0.0),
                    "recommendedSize": str(line.get("recommendedSize", "") or ""),
                    "capacityCfh": self._to_non_negative_number(line.get("capacityCfh", 0), 0.0),
                    "selectedLength": self._to_non_negative_number(line.get("selectedLength", 0), 0.0),
                    "maxAllowableLength": self._to_non_negative_number(line.get("maxAllowableLength", 0), 0.0),
                    "warning": str(line.get("warning", "") or ""),
                }
            )

        return {
            "schema_version": self._PROJECT_SCHEMA_VERSION,
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "activeModule": active_module,
            "project": {
                "activeProject": normalized_active_project,
                "savedProjects": normalized_saved_projects,
            },
            "settings": {
                "codeBasis": code_basis,
                "slope": slope,
                "zipCode": zip_code,
                "rainfallRate": rainfall_rate,
                "drainType": drain_type,
            },
            "areas": normalized_areas,
            "condensate": {
                "selection": {
                    "sizingMode": "TOTAL_TONS" if str(condensate_selection_raw.get("sizingMode", "PER_UNIT_QTY")) == "TOTAL_TONS" else "PER_UNIT_QTY",
                    "equipment": str(condensate_selection_raw.get("equipment", "AHU") or "AHU"),
                    "tonsPerUnit": self._to_non_negative_number(condensate_selection_raw.get("tonsPerUnit", 0.0), 0.0),
                    "quantity": self._to_non_negative_number(condensate_selection_raw.get("quantity", 1.0), 1.0),
                    "totalTons": self._to_non_negative_number(condensate_selection_raw.get("totalTons", 0.0), 0.0),
                    "equipmentRows": normalized_condensate_rows,
                    "zones": normalized_condensate_zones,
                },
                "result": {
                    "tableName": str(condensate_result_raw.get("tableName", "TABLE 814.3") or "TABLE 814.3"),
                    "sizingMode": "TOTAL_TONS" if str(condensate_result_raw.get("sizingMode", "PER_UNIT_QTY")) == "TOTAL_TONS" else "PER_UNIT_QTY",
                    "equipment": str(condensate_result_raw.get("equipment", "AHU") or "AHU"),
                    "equipmentLabel": str(condensate_result_raw.get("equipmentLabel", "Air Handler Unit (AHU)") or "Air Handler Unit (AHU)"),
                    "tonsPerUnit": self._to_non_negative_number(condensate_result_raw.get("tonsPerUnit", 0.0), 0.0),
                    "quantity": self._to_non_negative_number(condensate_result_raw.get("quantity", 0.0), 0.0),
                    "totalTonsInput": self._to_non_negative_number(condensate_result_raw.get("totalTonsInput", 0.0), 0.0),
                    "totalTonsUsed": self._to_non_negative_number(condensate_result_raw.get("totalTonsUsed", 0.0), 0.0),
                    "recommendedSize": str(condensate_result_raw.get("recommendedSize", "") or ""),
                    "warning": str(condensate_result_raw.get("warning", "") or ""),
                    "rows": normalized_condensate_result_rows,
                    "zones": normalized_condensate_result_zones,
                },
            },
            "vent": {
                "zones": normalized_vent_zones,
                "selection": {
                    "codeBasis": "CPC"
                    if str(vent_selection_raw.get("codeBasis", "IPC")) == "CPC"
                    else "IPC",
                    "usageType": _normalize_fixture_category(vent_selection_raw.get("usageType", "PRIVATE")),
                    "drainageOrientation": "VERTICAL"
                    if str(vent_selection_raw.get("drainageOrientation", "HORIZONTAL")) == "VERTICAL"
                    else "HORIZONTAL",
                    "useManualTotal": self._as_bool(vent_selection_raw.get("useManualTotal", False), False),
                    "autoTotalDfu": self._to_non_negative_number(vent_selection_raw.get("autoTotalDfu", 0.0), 0.0),
                    "manualTotalDfu": self._to_non_negative_number(vent_selection_raw.get("manualTotalDfu", 0.0), 0.0),
                    "totalDfu": self._to_non_negative_number(vent_selection_raw.get("totalDfu", 0.0), 0.0),
                    "fixtureRows": normalized_vent_rows,
                    "zones": normalized_vent_zones,
                },
                "result": {
                    "section": "vent",
                    "codeBasis": "CPC"
                    if str(vent_result_raw.get("codeBasis", "IPC")) == "CPC"
                    else "IPC",
                    "usageType": _normalize_fixture_category(
                        vent_result_raw.get("usageType", vent_selection_raw.get("usageType", "PRIVATE"))
                    ),
                    "drainageOrientation": "VERTICAL"
                    if str(vent_result_raw.get("drainageOrientation", vent_selection_raw.get("drainageOrientation", "HORIZONTAL"))) == "VERTICAL"
                    else "HORIZONTAL",
                    "dfu": self._to_non_negative_number(vent_result_raw.get("dfu", 0.0), 0.0),
                    "recommendedSize": str(vent_result_raw.get("recommendedSize", "") or ""),
                    "sizingTable": str(vent_result_raw.get("sizingTable", "TABLE 703.2") or "TABLE 703.2"),
                    "tableMaxUnits": self._to_non_negative_number(vent_result_raw.get("tableMaxUnits", 0.0), 0.0),
                    "maxLengthReference": str(vent_result_raw.get("maxLengthReference", "") or ""),
                    "warning": str(vent_result_raw.get("warning", "") or ""),
                    "autoTotalDfu": self._to_non_negative_number(vent_result_raw.get("autoTotalDfu", 0.0), 0.0),
                    "manualTotalDfu": self._to_non_negative_number(vent_result_raw.get("manualTotalDfu", 0.0), 0.0),
                    "useManualTotal": self._as_bool(vent_result_raw.get("useManualTotal", False), False),
                    "rows": [
                        row for row in vent_result_raw.get("rows", []) if isinstance(row, dict)
                    ],
                    "zones": [zone for zone in vent_result_raw.get("zones", []) if isinstance(zone, dict)],
                },
            },
            "sanitary_drainage": {
                "zones": normalized_vent_zones,
            },
            "ventilation": {
                "selection": {
                    "projectName": str(ventilation_selection_raw.get("projectName", "") or ""),
                    "systemType": str(ventilation_selection_raw.get("systemType", "SINGLE_ZONE") or "SINGLE_ZONE"),
                    "systemPopulationPs": self._to_non_negative_number(ventilation_selection_raw.get("systemPopulationPs", 0.0), 0.0),
                    "occupancySearch": str(ventilation_selection_raw.get("occupancySearch", "") or ""),
                    "zones": normalized_ventilation_zones,
                },
                "result": {
                    "section": "ventilation",
                    "status": str(ventilation_result_raw.get("status", "incomplete") or "incomplete"),
                    "message": str(ventilation_result_raw.get("message", "") or ""),
                    "systemType": str(ventilation_result_raw.get("systemType", "SINGLE_ZONE") or "SINGLE_ZONE"),
                    "sumPz": self._to_non_negative_number(ventilation_result_raw.get("sumPz", 0.0), 0.0),
                    "sumVoz": self._to_non_negative_number(ventilation_result_raw.get("sumVoz", 0.0), 0.0),
                    "ps": self._to_non_negative_number(ventilation_result_raw.get("ps", 0.0), 0.0),
                    "D": self._to_non_negative_number(ventilation_result_raw.get("D", 0.0), 0.0),
                    "Ev": self._to_non_negative_number(ventilation_result_raw.get("Ev", 0.0), 0.0),
                    "Vou": self._to_non_negative_number(ventilation_result_raw.get("Vou", 0.0), 0.0),
                    "Vot": self._to_non_negative_number(ventilation_result_raw.get("Vot", 0.0), 0.0),
                    "zones": [zone for zone in ventilation_result_raw.get("zones", []) if isinstance(zone, dict)],
                    "warnings": [str(w or "") for w in ventilation_result_raw.get("warnings", []) if str(w or "").strip()],
                },
            },
            "gas": {
                "tables": normalized_tables,
                "selection": {
                    "codeBasis": _normalize_code_basis(gas_selection_raw.get("codeBasis", "IPC")),
                    "gasType": _normalize_fuel(gas_selection_raw.get("gasType", "NATURAL_GAS")),
                    "inletPressureOption": "LT2"
                    if str(gas_selection_raw.get("inletPressureOption", "") or "").upper() == "LT2"
                    else "",
                    "inletPressure": self._to_non_negative_number(gas_selection_raw.get("inletPressure", 0.0), 0.0),
                    "pressureDrop": self._to_non_negative_number(gas_selection_raw.get("pressureDrop", 0.0), 0.0),
                    "specificGravity": self._to_non_negative_number(gas_selection_raw.get("specificGravity", 0.6), 0.6),
                    "material": _normalize_material(gas_selection_raw.get("material", "SCHEDULE_40_METALLIC")),
                    "demandMode": "BTUH" if str(gas_selection_raw.get("demandMode", "CFH")) == "BTUH" else "CFH",
                    "heatingValue": self._to_non_negative_number(gas_selection_raw.get("heatingValue", 1000.0), 1000.0),
                },
                "lines": normalized_lines,
                "result": {
                    "tableId": str(gas_result_raw.get("tableId", "") or ""),
                    "tableName": str(gas_result_raw.get("tableName", "") or ""),
                    "matchType": str(gas_result_raw.get("matchType", "") or ""),
                    "warning": str(gas_result_raw.get("warning", "") or ""),
                    "lines": normalized_result_lines,
                },
            },
            "wsfu": {
                "zones": normalized_wsfu_zones,
                "selection": {
                    "totalFu": self._to_non_negative_number(wsfu_selection_raw.get("totalFu", 0.0), 0.0),
                    "flushType": "FLUSH_VALVE"
                    if str(wsfu_selection_raw.get("flushType", "FLUSH_TANK")) == "FLUSH_VALVE"
                    else "FLUSH_TANK",
                    "designLengthFt": self._to_non_negative_number(wsfu_selection_raw.get("designLengthFt", 100.0), 100.0),
                    "usageType": _normalize_fixture_category(wsfu_selection_raw.get("usageType", "PRIVATE")),
                    "useManualTotal": self._as_bool(wsfu_selection_raw.get("useManualTotal", False), False),
                    "manualTotalFu": self._to_non_negative_number(wsfu_selection_raw.get("manualTotalFu", 0.0), 0.0),
                    "autoTotalFu": self._to_non_negative_number(wsfu_selection_raw.get("autoTotalFu", 0.0), 0.0),
                    "fixtureRows": normalized_wsfu_rows,
                    "zones": normalized_wsfu_zones,
                },
                "result": {
                    "totalFu": self._to_non_negative_number(wsfu_result_raw.get("totalFu", 0.0), 0.0),
                    "flushTankGpm": self._to_non_negative_number(wsfu_result_raw.get("flushTankGpm", 0.0), 0.0),
                    "flushValveGpm": self._to_non_negative_number(wsfu_result_raw.get("flushValveGpm", 0.0), 0.0),
                    "matchedTankFu": self._to_non_negative_number(wsfu_result_raw.get("matchedTankFu", 0.0), 0.0),
                    "matchedValveFu": self._to_non_negative_number(wsfu_result_raw.get("matchedValveFu", 0.0), 0.0),
                    "flushType": "FLUSH_VALVE"
                    if str(wsfu_result_raw.get("flushType", "FLUSH_TANK")) == "FLUSH_VALVE"
                    else "FLUSH_TANK",
                    "selectedFlowGpm": self._to_non_negative_number(wsfu_result_raw.get("selectedFlowGpm", 0.0), 0.0),
                    "designLengthFt": self._to_non_negative_number(wsfu_result_raw.get("designLengthFt", 100.0), 100.0),
                    "serviceSize": str(wsfu_result_raw.get("serviceSize", "") or ""),
                    "sizeCapacityGpm": self._to_non_negative_number(wsfu_result_raw.get("sizeCapacityGpm", 0.0), 0.0),
                    "frictionLossByPipeType": [
                        row
                        for row in wsfu_result_raw.get("frictionLossByPipeType", [])
                        if isinstance(row, dict)
                    ],
                    "source": str(wsfu_result_raw.get("source", "") or ""),
                    "zones": [zone for zone in wsfu_result_raw.get("zones", []) if isinstance(zone, dict)],
                },
            },
            "fixtureUnit": {
                "selection": {
                    "codeBasis": "CPC" if str(fixture_unit_selection_raw.get("codeBasis", "IPC")) == "CPC" else "IPC",
                    "drainageOrientation": "VERTICAL"
                    if str(fixture_unit_selection_raw.get("drainageOrientation", "HORIZONTAL")) == "VERTICAL"
                    else "HORIZONTAL",
                    "flushType": "FLUSH_VALVE"
                    if str(fixture_unit_selection_raw.get("flushType", "FLUSH_TANK")) == "FLUSH_VALVE"
                    else "FLUSH_TANK",
                    "designLengthFt": self._to_non_negative_number(fixture_unit_selection_raw.get("designLengthFt", 100.0), 100.0),
                    "zones": normalized_fixture_unit_zones,
                    "rows": normalized_fixture_unit_rows,
                },
                "result": fixture_unit_result_raw,
            },
            "duct": {
                "zones": normalized_duct_zones,
                "selection": {
                    "mode": str(duct_selection_raw.get("mode", "SIZE_FROM_CFM_FRICTION") or "SIZE_FROM_CFM_FRICTION"),
                    "shape": str(duct_selection_raw.get("shape", "ROUND") or "ROUND"),
                    "cfm": self._to_non_negative_number(duct_selection_raw.get("cfm", 0.0), 0.0),
                    "frictionTarget": self._to_non_negative_number(duct_selection_raw.get("frictionTarget", 0.08), 0.08),
                    "diameter": self._to_non_negative_number(duct_selection_raw.get("diameter", 12.0), 12.0),
                    "width": self._to_non_negative_number(duct_selection_raw.get("width", 12.0), 12.0),
                    "height": self._to_non_negative_number(duct_selection_raw.get("height", 10.0), 10.0),
                    "ratioLimit": self._to_non_negative_number(duct_selection_raw.get("ratioLimit", 3.0), 3.0),
                    "zones": normalized_duct_zones,
                },
                "result": {
                    "status": str(duct_result_raw.get("status", "") or ""),
                    "message": str(duct_result_raw.get("message", "") or ""),
                    "equivalentDiameter": self._to_non_negative_number(duct_result_raw.get("equivalentDiameter", 0.0), 0.0),
                    "recommendedRound": str(duct_result_raw.get("recommendedRound", "") or ""),
                    "recommendedRect": str(duct_result_raw.get("recommendedRect", "") or ""),
                    "velocityFpm": self._to_non_negative_number(duct_result_raw.get("velocityFpm", 0.0), 0.0),
                    "velocityPressure": self._to_non_negative_number(duct_result_raw.get("velocityPressure", 0.0), 0.0),
                    "frictionRate": self._to_non_negative_number(duct_result_raw.get("frictionRate", 0.0), 0.0),
                    "warning": str(duct_result_raw.get("warning", "") or ""),
                    "zones": [zone for zone in duct_result_raw.get("zones", []) if isinstance(zone, dict)],
                },
            },
            "duct_sizing": {
                "zones": normalized_duct_zones,
            },
            "ductStatic": {
                "selection": {
                    "segments": normalized_duct_static_segments,
                },
                "result": {
                    "segments": [
                        seg for seg in duct_static_result_raw.get("segments", []) if isinstance(seg, dict)
                    ],
                    "totalPressureDrop": self._to_non_negative_number(
                        duct_static_result_raw.get("totalPressureDrop", 0.0), 0.0
                    ),
                    "warnings": [
                        str(w or "") for w in duct_static_result_raw.get("warnings", []) if str(w or "").strip()
                    ],
                },
            },
            "solar": {
                "selection": {
                    "buildingSf": self._to_non_negative_number(solar_selection_raw.get("buildingSf", 0.0), 0.0),
                    "zipCode": self._normalize_zip(str(solar_selection_raw.get("zipCode", "") or "")),
                    "climateZone": self._to_non_negative_number(solar_selection_raw.get("climateZone", 0.0), 0.0),
                    "buildingType": str(solar_selection_raw.get("buildingType", "Grocery") or "Grocery"),
                    "useSaraPath": self._as_bool(solar_selection_raw.get("useSaraPath", False), False),
                    "sara": self._to_non_negative_number(solar_selection_raw.get("sara", 0.0), 0.0),
                    "dValue": self._to_non_negative_number(solar_selection_raw.get("dValue", 0.95), 0.95),
                    "solarExempt": self._as_bool(solar_selection_raw.get("solarExempt", False), False),
                    "batteryExempt": self._as_bool(solar_selection_raw.get("batteryExempt", False), False),
                    "notes": str(solar_selection_raw.get("notes", "") or ""),
                },
                "result": {
                    "status": str(solar_result_raw.get("status", "pending_workbook") or "pending_workbook"),
                    "message": str(
                        solar_result_raw.get(
                            "message",
                            "Solar workbook source is pending. Provide local .xls/.xlsx path to activate formulas.",
                        )
                        or "Solar workbook source is pending. Provide local .xls/.xlsx path to activate formulas."
                    ),
                    "climateGroup": str(solar_result_raw.get("climateGroup", "") or ""),
                    "factorA": self._to_non_negative_number(solar_result_raw.get("factorA", 0.0), 0.0),
                    "factorB": self._to_non_negative_number(solar_result_raw.get("factorB", 0.0), 0.0),
                    "factorC": self._to_non_negative_number(solar_result_raw.get("factorC", 0.0), 0.0),
                    "minimumSara": self._to_non_negative_number(solar_result_raw.get("minimumSara", 0.0), 0.0),
                    "saraUsed": self._to_non_negative_number(solar_result_raw.get("saraUsed", 0.0), 0.0),
                    "pvKwByCfa": self._to_non_negative_number(solar_result_raw.get("pvKwByCfa", 0.0), 0.0),
                    "pvKwBySara": self._to_non_negative_number(solar_result_raw.get("pvKwBySara", 0.0), 0.0),
                    "pvKwdc": self._to_non_negative_number(solar_result_raw.get("pvKwdc", 0.0), 0.0),
                    "batteryKwh": self._to_non_negative_number(solar_result_raw.get("batteryKwh", 0.0), 0.0),
                    "batteryKw": self._to_non_negative_number(solar_result_raw.get("batteryKw", 0.0), 0.0),
                    "notes": str(solar_result_raw.get("notes", "") or ""),
                },
            },
            "refrigerant": {
                "selection": refrigerant_selection_raw,
                "result": refrigerant_result_raw,
            },
        }

    @staticmethod
    def _extract_pdf_text(pdf_path: Path) -> str:
        try:
            from pypdf import PdfReader  # type: ignore
        except Exception:
            raise RuntimeError(
                "PDF parser dependency is missing. Install with: pip install pypdf"
            )

        reader = PdfReader(str(pdf_path))
        page_text = []
        for page in reader.pages:
            try:
                page_text.append(page.extract_text() or "")
            except Exception:
                page_text.append("")
        return "\n".join(page_text)

    @staticmethod
    def _detect_gas_type(text: str) -> str:
        t = text.lower()
        if "propane" in t or "lp gas" in t or "liquefied petroleum" in t:
            return "PROPANE"
        return "NATURAL_GAS"

    @staticmethod
    def _detect_code_basis(text: str) -> str:
        t = text.lower()
        if "california plumbing code" in t or "cpc" in t:
            return "CPC"
        return "IPC"

    @staticmethod
    def _fixture_category_label(raw) -> str:
        value = str(raw or "PRIVATE").upper()
        if value == "ASSEMBLY":
            return "Assembly"
        if value == "PUBLIC":
            return "Public"
        return "Private"

    @staticmethod
    def _parse_gas_table_drafts(text: str, source_name: str) -> list:
        lines = [ln.strip() for ln in text.splitlines() if ln and ln.strip()]
        if not lines:
            return []

        candidate_sizes = re.findall(r'(?:\d+-\d+/\d+|\d+/\d+|\d+)"', text)
        seen = []
        for s in candidate_sizes:
            if s not in seen:
                seen.append(s)
        sizes = seen if len(seen) >= 3 else ['1/2"', '3/4"', '1"', '1-1/4"', '1-1/2"', '2"', '2-1/2"', '3"']

        rows = []
        row_pattern = re.compile(r'^\s*(\d{1,4})\s+([\d,\s]{8,})$')
        for ln in lines:
            m = row_pattern.match(ln)
            if not m:
                continue
            length = float(m.group(1))
            numbers = [float(n.replace(",", "")) for n in re.findall(r'\d[\d,]*', m.group(2))]
            if len(numbers) < 3:
                continue
            caps = {}
            max_cols = min(len(sizes), len(numbers))
            for idx in range(max_cols):
                caps[sizes[idx]] = numbers[idx]
            if caps:
                rows.append({"length": length, "capacities": caps})

        inlet_match = re.search(r'inlet pressure[^0-9]*(\d+(?:\.\d+)?)', text, flags=re.IGNORECASE)
        drop_match = re.search(r'pressure drop[^0-9]*(\d+(?:\.\d+)?)', text, flags=re.IGNORECASE)
        sg_match = re.search(r'specific gravity[^0-9]*(\d+(?:\.\d+)?)', text, flags=re.IGNORECASE)

        draft = {
            "draftId": hashlib.sha1(f"{source_name}-{len(text)}".encode("utf-8")).hexdigest()[:12],
            "sourceFile": source_name,
            "tableName": lines[0][:120] if lines else "Gas Table Draft",
            "codeBasis": Api._detect_code_basis(text),
            "gasType": Api._detect_gas_type(text),
            "inletPressure": float(inlet_match.group(1)) if inlet_match else 0.0,
            "pressureDrop": float(drop_match.group(1)) if drop_match else 0.0,
            "specificGravity": float(sg_match.group(1)) if sg_match else 0.6,
            "material": "SCHEDULE_40_METALLIC",
            "rows": rows,
            "rawPreview": "\n".join(lines[:80]),
        }
        return [draft]

    def upload_gas_table_pdfs_legacy(self) -> dict:
        try:
            window = self._active_window()
            if window is None:
                return {"ok": False, "reason": "App window is not ready yet."}

            dialog_result = window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=True,
                file_types=("PDF Files (*.pdf)",),
            )
            selected_paths = []
            if isinstance(dialog_result, (list, tuple)):
                selected_paths = [str(p) for p in dialog_result]
            elif dialog_result:
                selected_paths = [str(dialog_result)]
            if not selected_paths:
                return {"ok": False, "reason": "No PDF files selected."}

            files = []
            warnings = []
            for path_str in selected_paths:
                pdf_path = Path(path_str)
                if not pdf_path.exists():
                    warnings.append(f"File not found: {pdf_path}")
                    continue
                if pdf_path.suffix.lower() != ".pdf":
                    warnings.append(f"Skipped non-PDF file: {pdf_path.name}")
                    continue
                text = self._extract_pdf_text(pdf_path)
                drafts = self._parse_gas_table_drafts(text, pdf_path.name)
                files.append(
                    {
                        "path": str(pdf_path),
                        "sourceFile": pdf_path.name,
                        "drafts": drafts,
                    }
                )
            if not files:
                return {"ok": False, "reason": "No valid PDF tables were extracted.", "warnings": warnings}
            return {"ok": True, "files": files, "warnings": warnings}
        except Exception as exc:
            return {"ok": False, "reason": f"Gas table upload failed: {exc}"}

    def _save_project_xlsx(self, target_path: Path, payload: dict) -> None:
        if target_path.exists():
            workbook = load_workbook(target_path)
        else:
            workbook = Workbook()

        if self._PROJECT_SHEET in workbook.sheetnames:
            sheet = workbook[self._PROJECT_SHEET]
        else:
            sheet = workbook.create_sheet(self._PROJECT_SHEET)

        sheet.sheet_state = "hidden"
        sheet["A1"] = "schema_version"
        sheet["B1"] = payload.get("schema_version", self._PROJECT_SCHEMA_VERSION)
        sheet["A2"] = "saved_at"
        sheet["B2"] = payload.get("saved_at", datetime.now().isoformat(timespec="seconds"))
        sheet["A3"] = "payload_json"
        sheet["B3"] = json.dumps(payload, ensure_ascii=False)
        workbook.save(target_path)

    def _load_project_xlsx(self, source_path: Path) -> dict:
        workbook = load_workbook(source_path, data_only=True)
        if self._PROJECT_SHEET not in workbook.sheetnames:
            raise ValueError("Project data sheet was not found in this workbook.")
        sheet = workbook[self._PROJECT_SHEET]
        raw_payload = sheet["B3"].value
        if raw_payload is None or str(raw_payload).strip() == "":
            raise ValueError("Project payload is missing in this workbook.")
        try:
            parsed = json.loads(str(raw_payload))
        except Exception:
            raise ValueError("Project payload in workbook is not valid JSON.")
        return self._normalize_project_payload(parsed)

    def _load_lookup_data_legacy(self) -> None:
        try:
            zip_rain_path = self._lookup_file_path("USA_ZIP_100yr_60min_rainfall_corrected.csv")

            if not zip_rain_path.exists():
                self._data_error = f"ZIP rainfall dataset not found: {zip_rain_path}"
                return

            # ZIP index with direct rainfall (no station matching required).
            with zip_rain_path.open(newline="", encoding="utf-8-sig") as f_zip:
                zip_reader = csv.DictReader(f_zip)
                for row in zip_reader:
                    z = self._normalize_zip(row.get("zip", ""))
                    lat = self._to_float(row.get("lat", ""))
                    lng = self._to_float(row.get("lng", ""))
                    rainfall_raw = self._to_float(row.get("rainfall_100yr_60min_in_hr", ""))
                    if not z or lat is None or lng is None or rainfall_raw is None:
                        continue
                    self._zip_index[z] = {
                        "lat": lat,
                        "lng": lng,
                        "city": str(row.get("city", "")).strip(),
                        "state_id": str(row.get("state_id", "")).strip(),
                        "state_name": str(row.get("state_name", "")).strip(),
                        "county_name": str(row.get("county_name", "")).strip(),
                        "rainfall_source_in_per_hr": rainfall_raw,
                        "rainfall_in_per_hr": self._round_up_one_decimal(rainfall_raw),
                    }

            if not self._zip_index:
                self._data_error = "ZIP rainfall dataset loaded, but no valid rows were found."
        except Exception as exc:
            self._data_error = f"Lookup data load failed: {exc}"

    def _load_wsfu_data_legacy(self) -> None:
        try:
            wsfu_path = self._lookup_file_path("WSFU2GPM.xlsx")
            if not wsfu_path.exists():
                self._wsfu_error = f"WSFU workbook not found: {wsfu_path}"
                return

            workbook = load_workbook(wsfu_path, data_only=True)
            if "FU" not in workbook.sheetnames:
                self._wsfu_error = "WSFU workbook is missing sheet 'FU'."
                return
            sheet = workbook["FU"]

            tank_rows = []
            valve_rows = []
            for row in sheet.iter_rows(min_row=3, max_col=4, values_only=True):
                fu_tank = self._to_float(row[0])
                gpm_tank = self._to_float(row[1])
                fu_valve = self._to_float(row[2])
                gpm_valve = self._to_float(row[3])

                if fu_tank is not None and gpm_tank is not None:
                    tank_rows.append((fu_tank, gpm_tank))
                if fu_valve is not None and gpm_valve is not None:
                    valve_rows.append((fu_valve, gpm_valve))

            if not tank_rows or not valve_rows:
                self._wsfu_error = "WSFU workbook contains no usable lookup rows."
                return

            # Ensure VLOOKUP-TRUE semantics can be applied safely.
            tank_rows.sort(key=lambda x: x[0])
            valve_rows.sort(key=lambda x: x[0])
            self._wsfu_tank_curve = tank_rows
            self._wsfu_valve_curve = valve_rows
            self._wsfu_error = ""
        except Exception as exc:
            self._wsfu_error = f"WSFU table load failed: {exc}"

    def _load_solar_zip_data_legacy(self) -> None:
        try:
            self._solar_zip_to_cz = {}
            path = self._lookup_file_path("solar_zip_climate.csv")
            if not path.exists():
                return
            with path.open(newline="", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    z = self._normalize_zip(row.get("zip", ""))
                    try:
                        cz = int(str(row.get("climate_zone", "")).strip())
                    except Exception:
                        continue
                    if re.fullmatch(r"\d{5}", z):
                        self._solar_zip_to_cz[z] = cz
        except Exception:
            self._solar_zip_to_cz = {}

    @staticmethod
    def _solar_climate_group(climate_zone: int) -> str:
        if climate_zone == 15:
            return "15"
        if climate_zone in (1, 3, 5):
            return "1,3,5,6"
        return "2,4,6-14"

    @staticmethod
    def _as_bool(value, default=False) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        text = str(value).strip().lower()
        if text in ("1", "true", "yes", "y", "on"):
            return True
        if text in ("0", "false", "no", "n", "off"):
            return False
        return default

    @staticmethod
    def _vlookup_true(value: float, curve: list) -> tuple:
        if not curve:
            return (0.0, 0.0)
        keys = [row[0] for row in curve]
        idx = bisect_right(keys, value) - 1
        if idx < 0:
            idx = 0
        if idx >= len(curve):
            idx = len(curve) - 1
        return curve[idx]

    def _recommend_wsfu_service_size(self, flow_gpm: float) -> tuple:
        for row in self._WSFU_SIZE_MAX_GPM:
            if flow_gpm <= row["maxGpm"]:
                return row["size"], row["maxGpm"], row["diameterIn"]
        last = self._WSFU_SIZE_MAX_GPM[-1]
        return f'>{last["size"]}', last["maxGpm"], last["diameterIn"]

    @staticmethod
    def _hazen_williams_head_loss_per_100ft(flow_gpm: float, c_factor: float, diameter_in: float) -> float:
        if flow_gpm <= 0 or c_factor <= 0 or diameter_in <= 0:
            return 0.0
        return 4.52 * (flow_gpm ** 1.85) / ((c_factor ** 1.85) * (diameter_in ** 4.87))

    def lookup_zip_rainfall_legacy(self, zip_code: str) -> dict:
        if self._data_error:
            return {"ok": False, "reason": self._data_error}

        z = self._normalize_zip(zip_code)
        if not re.fullmatch(r"\d{5}", z):
            return {"ok": False, "reason": "Enter a valid 5-digit ZIP code."}

        zip_entry = self._zip_index.get(z)
        if not zip_entry:
            return {"ok": False, "reason": f"ZIP code {z} was not found in the ZIP dataset."}

        return {
            "ok": True,
            "zip": z,
            "city": zip_entry["city"],
            "state_id": zip_entry["state_id"],
            "state_name": zip_entry["state_name"],
            "county_name": zip_entry["county_name"],
            "latitude": zip_entry["lat"],
            "longitude": zip_entry["lng"],
            "rainfall_source_in_per_hr": zip_entry["rainfall_source_in_per_hr"],
            "rainfall_in_per_hr": zip_entry["rainfall_in_per_hr"],
        }

    def calculate_wsfu_legacy(self, payload) -> dict:
        if self._wsfu_error:
            return {"ok": False, "reason": self._wsfu_error}

        if isinstance(payload, dict):
            fu = self._to_non_negative_number(payload.get("totalFu", 0), -1.0)
            flush_type = str(payload.get("flushType", "FLUSH_TANK") or "FLUSH_TANK").upper()
            design_length_ft = self._to_non_negative_number(payload.get("designLengthFt", 100), 100.0)
        else:
            fu = self._to_non_negative_number(payload, -1.0)
            flush_type = "FLUSH_TANK"
            design_length_ft = 100.0

        if fu < 0:
            return {"ok": False, "reason": "Enter a non-negative fixture unit value."}
        if flush_type not in self._WSFU_FLUSH_TYPES:
            flush_type = "FLUSH_TANK"

        if not self._wsfu_tank_curve or not self._wsfu_valve_curve:
            return {"ok": False, "reason": "WSFU lookup tables are not available."}

        tank_match_fu, tank_gpm = self._vlookup_true(fu, self._wsfu_tank_curve)
        valve_match_fu, valve_gpm = self._vlookup_true(fu, self._wsfu_valve_curve)
        selected_flow_gpm = valve_gpm if flush_type == "FLUSH_VALVE" else tank_gpm
        service_size, max_gpm, diameter_in = self._recommend_wsfu_service_size(selected_flow_gpm)

        friction_rows = []
        for pipe_type in self._WSFU_PIPE_TYPE_C_FACTORS:
            c_factor = pipe_type["cFactor"]
            head_loss_per_100 = self._hazen_williams_head_loss_per_100ft(selected_flow_gpm, c_factor, diameter_in)
            psi_loss_per_100 = head_loss_per_100 / 2.31
            psi_loss_for_length = psi_loss_per_100 * (design_length_ft / 100.0)
            friction_rows.append(
                {
                    "pipeType": pipe_type["label"],
                    "cFactor": c_factor,
                    "headLossFtPer100": round(head_loss_per_100, 3),
                    "psiLossPer100": round(psi_loss_per_100, 3),
                    "psiLossForLength": round(psi_loss_for_length, 3),
                }
            )

        warning = ""
        if service_size.startswith(">"):
            warning = "Flow exceeds built-in size range. Increase service size and verify with code table."

        return {
            "ok": True,
            "totalFu": fu,
            "flushTankGpm": tank_gpm,
            "flushValveGpm": valve_gpm,
            "matchedTankFu": tank_match_fu,
            "matchedValveFu": valve_match_fu,
            "flushType": flush_type,
            "selectedFlowGpm": selected_flow_gpm,
            "designLengthFt": design_length_ft,
            "serviceSize": service_size,
            "sizeCapacityGpm": max_gpm,
            "frictionLossByPipeType": friction_rows,
            "warning": warning,
            "source": "WSFU2GPM.xlsx lookup + Hazen-Williams friction by selected sizing basis",
        }

    def get_solar_catalog_legacy(self) -> dict:
        try:
            return {
                "ok": True,
                "buildingTypes": list(self._SOLAR_BUILDING_TYPES),
                "climateGroups": list(self._SOLAR_CLIMATE_GROUPS),
                "zipCount": len(self._solar_zip_to_cz),
            }
        except Exception as exc:
            return {"ok": False, "reason": f"Solar catalog failed: {exc}"}

    def lookup_solar_climate_legacy(self, zip_code: str) -> dict:
        try:
            z = self._normalize_zip(str(zip_code or ""))
            if not re.fullmatch(r"\d{5}", z or ""):
                return {"ok": False, "reason": "Enter a valid 5-digit ZIP code."}
            climate_zone = int(self._solar_zip_to_cz.get(z, 0))
            if climate_zone <= 0:
                return {"ok": False, "reason": "No climate mapping found for ZIP."}
            return {"ok": True, "zipCode": z, "climateZone": climate_zone}
        except Exception as exc:
            return {"ok": False, "reason": f"Solar climate lookup failed: {exc}"}

    def calculate_solar_legacy(self, payload: dict) -> dict:
        try:
            if not isinstance(payload, dict):
                payload = {}

            building_sf = self._to_non_negative_number(payload.get("buildingSf", 0), 0.0)
            zip_code = self._normalize_zip(str(payload.get("zipCode", "") or ""))
            climate_zone = int(self._to_non_negative_number(payload.get("climateZone", 0), 0))
            if climate_zone <= 0 and re.fullmatch(r"\d{5}", zip_code or ""):
                climate_zone = int(self._solar_zip_to_cz.get(zip_code, 0))
            if climate_zone <= 0:
                return {"ok": False, "reason": "Enter a valid ZIP with climate mapping or a climate zone."}

            building_type = str(payload.get("buildingType", "Grocery") or "Grocery")
            if building_type not in self._SOLAR_BUILDING_TYPES:
                building_type = "Grocery"

            use_sara_path = self._as_bool(payload.get("useSaraPath", False), False)
            sara_input = self._to_non_negative_number(payload.get("sara", 0), 0.0)
            d_value = self._to_non_negative_number(payload.get("dValue", 0.95), 0.95)
            if d_value <= 0:
                d_value = 0.95

            solar_exempt = self._as_bool(payload.get("solarExempt", False), False)
            battery_exempt = self._as_bool(payload.get("batteryExempt", False), False)

            factors = self._default_solar_factor_matrix()
            climate_group = self._solar_climate_group(climate_zone)
            factor_row = factors.get(building_type, {}).get(climate_group, {"A": 0.44, "B": 0.93, "C": 0.23})
            factor_a = self._to_non_negative_number(payload.get("factorA", factor_row.get("A", 0.44)), factor_row.get("A", 0.44))
            factor_b = self._to_non_negative_number(payload.get("factorB", factor_row.get("B", 0.93)), factor_row.get("B", 0.93))
            factor_c = self._to_non_negative_number(payload.get("factorC", factor_row.get("C", 0.23)), factor_row.get("C", 0.23))

            minimum_sara = building_sf * 0.15
            sara_used = sara_input if (use_sara_path and sara_input > 0) else minimum_sara

            pv_kw_by_cfa = (building_sf * factor_a) / 1000.0
            pv_kw_by_sara = (sara_used * 14.0) / 1000.0
            pv_kwdc = min(pv_kw_by_cfa, pv_kw_by_sara) if pv_kw_by_cfa > 0 and pv_kw_by_sara > 0 else max(pv_kw_by_cfa, pv_kw_by_sara)
            battery_kwh = (pv_kwdc * factor_b) / math.sqrt(d_value) if d_value > 0 else 0.0
            battery_kw = pv_kwdc * factor_c

            solar_status = "EXEMPT" if solar_exempt else "NOT EXEMPT"
            battery_status = "EXEMPT" if battery_exempt else "NOT EXEMPT"
            message = "Solar and battery outputs calculated from workbook-based equation flow."

            return {
                "ok": True,
                "status": "calculated",
                "message": message,
                "buildingSf": building_sf,
                "zipCode": zip_code,
                "climateZone": climate_zone,
                "climateGroup": climate_group,
                "buildingType": building_type,
                "useSaraPath": use_sara_path,
                "saraInput": sara_input,
                "minimumSara": minimum_sara,
                "saraUsed": sara_used,
                "dValue": d_value,
                "factorA": factor_a,
                "factorB": factor_b,
                "factorC": factor_c,
                "pvKwByCfa": pv_kw_by_cfa,
                "pvKwBySara": pv_kw_by_sara,
                "pvKwdc": pv_kwdc,
                "batteryKwh": battery_kwh,
                "batteryKw": battery_kw,
                "solarExemptionStatus": solar_status,
                "batteryExemptionStatus": battery_status,
            }
        except Exception as exc:
            return {"ok": False, "reason": f"Solar calculation failed: {exc}"}

    def save_project(self, payload: dict) -> dict:
        try:
            normalized = self._normalize_project_payload(payload)
            window = self._active_window()
            if window is None:
                return {"ok": False, "reason": "App window is not ready yet."}

            dialog_result = window.create_file_dialog(
                webview.SAVE_DIALOG,
                save_filename=f"MechanicalAndPlumbingCalculatorProject_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                file_types=("Project Files (*.json;*.xlsx)", "JSON Files (*.json)", "Excel Files (*.xlsx)"),
            )
            selected_path = self._dialog_path(dialog_result).strip()
            if not selected_path:
                return {"ok": False, "reason": "Save canceled by user."}

            target_path = Path(selected_path)
            suffix = target_path.suffix.lower()
            if suffix not in (".json", ".xlsx"):
                target_path = target_path.with_suffix(".json")
                suffix = ".json"

            if suffix == ".json":
                target_path.write_text(
                    json.dumps(normalized, indent=2, ensure_ascii=False),
                    encoding="utf-8",
                )
            else:
                self._save_project_xlsx(target_path, normalized)

            return {"ok": True, "path": str(target_path)}
        except Exception as exc:
            return {"ok": False, "reason": f"Save project failed: {exc}"}

    def load_project(self) -> dict:
        try:
            window = self._active_window()
            if window is None:
                return {"ok": False, "reason": "App window is not ready yet."}

            dialog_result = window.create_file_dialog(
                webview.OPEN_DIALOG,
                allow_multiple=False,
                file_types=("Project Files (*.json;*.xlsx)", "JSON Files (*.json)", "Excel Files (*.xlsx)"),
            )
            selected_path = self._dialog_path(dialog_result).strip()
            if not selected_path:
                return {"ok": False, "reason": "Load canceled by user."}

            source_path = Path(selected_path)
            suffix = source_path.suffix.lower()
            if suffix == ".json":
                parsed = json.loads(source_path.read_text(encoding="utf-8"))
                normalized = self._normalize_project_payload(parsed)
            elif suffix == ".xlsx":
                normalized = self._load_project_xlsx(source_path)
            else:
                return {"ok": False, "reason": "Unsupported file type. Use .json or .xlsx."}

            return {"ok": True, "path": str(source_path), "data": normalized}
        except Exception as exc:
            return {"ok": False, "reason": f"Load project failed: {exc}"}

    def export_per_area_legacy(self, snapshot: dict) -> dict:
        try:
            rows = snapshot.get("rows", [])
            if not rows:
                return {"ok": False, "reason": "No per-area rows available to export."}

            template_path = self._template_path()
            if not template_path.exists():
                return {"ok": False, "reason": f"Template file not found: {template_path}"}

            workbook = load_workbook(template_path)
            sheet = workbook.worksheets[0]

            drain_type_text = snapshot.get("drainTypeLabel", "")
            demand_multiplier = snapshot.get("demandMultiplier", "1.00x")
            if demand_multiplier != "1.00x":
                drain_type_text = f"{drain_type_text} ({demand_multiplier})"

            project_info = self._extract_export_project(snapshot)
            project_pairs = [
                ("Project Name", str(project_info.get("projectName", "") or "").strip()),
                ("Project Number", str(project_info.get("projectNumber", "") or "").strip()),
                ("Client", str(project_info.get("clientName", "") or "").strip()),
                ("Location", str(project_info.get("projectLocation", "") or "").strip()),
                ("Date", str(project_info.get("date", "") or "").strip()),
                ("Revision", str(project_info.get("revision", "") or "").strip()),
                ("Engineer", str(project_info.get("engineerName", "") or "").strip()),
                ("Company", str(project_info.get("companyName", "") or "").strip()),
            ]
            project_row = 2
            for label, value in project_pairs:
                if not value:
                    continue
                sheet.cell(row=project_row, column=6, value=label)
                sheet.cell(row=project_row, column=7, value=value)
                project_row += 1

            sheet.cell(row=2, column=2, value=snapshot.get("codeBasisLabel", ""))
            sheet.cell(row=3, column=1, value=snapshot.get("codeBasisLabel", ""))
            sheet.cell(row=3, column=2, value=snapshot.get("slopeLabel", ""))
            try:
                rainfall_rate = float(snapshot.get("rainfallRate", "0"))
            except Exception:
                rainfall_rate = snapshot.get("rainfallRate", "")
            sheet.cell(row=3, column=3, value=rainfall_rate)
            sheet.cell(row=3, column=4, value=drain_type_text)

            headers = [
                "Area",
                "Parent Area (sq ft)",
                "Sub-Areas (sq ft)",
                "Roof Area (sq ft)",
                "Side Wall Areas (sq ft)",
                "Side Wall Contribution (sq ft)",
                "Effective Drainage Area (sq ft)",
                "Effective Runoff (GPM)",
                "Effective Runoff (CFS)",
                "Horizontal Drain Size (in)",
                "Maximum Horizontal Capacity (GPM)",
                "Vertical Drain Size (in)",
                "Maximum Vertical Capacity (GPM)",
            ]
            for i, title in enumerate(headers, start=1):
                sheet.cell(row=5, column=i, value=title)

            max_clear_row = max(sheet.max_row + 10, 1000)
            for r in range(6, max_clear_row + 1):
                for c in range(1, 14):
                    sheet.cell(row=r, column=c, value=None)

            for i, row in enumerate(rows):
                r = 6 + i
                sheet.cell(row=r, column=1, value=row.get("label"))
                sheet.cell(row=r, column=2, value=row.get("parentArea"))
                sheet.cell(row=r, column=3, value=row.get("subAreaListText"))
                sheet.cell(row=r, column=4, value=row.get("roofArea"))
                sheet.cell(row=r, column=5, value=row.get("sideWallAreasText"))
                sheet.cell(row=r, column=6, value=row.get("sideWallContribution"))
                sheet.cell(row=r, column=7, value=row.get("area"))
                sheet.cell(row=r, column=8, value=row.get("effectiveGpm"))
                sheet.cell(row=r, column=9, value=row.get("effectiveCfs"))
                sheet.cell(row=r, column=10, value=row.get("horizontalSize"))
                sheet.cell(row=r, column=11, value=row.get("horizontalCapacity"))
                sheet.cell(row=r, column=12, value=row.get("verticalSize"))
                sheet.cell(row=r, column=13, value=row.get("verticalCapacity"))

            output_dir = Path.home() / "Downloads"
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"Per_Area_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = output_dir / filename
            workbook.save(output_path)

            return {
                "ok": True,
                "rowCount": len(rows),
                "exportedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "path": str(output_path),
            }
        except Exception as exc:
            return {"ok": False, "reason": f"Export failed: {exc}"}

    def export_section_legacy(self, snapshot: dict) -> dict:
        try:
            if not isinstance(snapshot, dict):
                return {"ok": False, "reason": "Invalid export payload."}

            section = str(snapshot.get("section", "") or "").strip().lower()
            if section == "storm":
                storm_snapshot = snapshot.get("storm", {})
                if not isinstance(storm_snapshot, dict):
                    return {"ok": False, "reason": "Invalid storm export payload."}
                return self.export_per_area(storm_snapshot)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Section Export"
            now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            section_label = section.title() if section else "Unknown"
            if section == "vent":
                section_label = "Sanitary Drainage"
            if section == "ventilation":
                section_label = "Ventilation Calculator"
            if section == "refrigerant":
                section_label = "Refrigerant Compliance"
            if section == "fixtureunit":
                section_label = "Fixture Unit Calculator"
            sheet.cell(row=1, column=1, value="Mechanical and Plumbing Calculator Export")
            sheet.cell(row=2, column=1, value="Section")
            sheet.cell(row=2, column=2, value=section_label)
            sheet.cell(row=3, column=1, value="Exported At")
            sheet.cell(row=3, column=2, value=now_text)
            row = 5
            row = self._write_project_header(sheet, row, self._extract_export_project(snapshot))
            if section == "home":
                summary = snapshot.get("summary", {})
                if not isinstance(summary, dict):
                    summary = {}
                sheet.cell(row=row, column=1, value="Active Module")
                sheet.cell(row=row, column=2, value=summary.get("activeModule", "home"))
                row += 1
                sheet.cell(row=row, column=1, value="Areas Count")
                sheet.cell(row=row, column=2, value=summary.get("areasCount", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Gas Lines Count")
                sheet.cell(row=row, column=2, value=summary.get("gasLinesCount", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Notes")
                sheet.cell(row=row, column=2, value="Home summary export")
                data_rows = 1
            elif section == "condensate":
                sheet.cell(row=row, column=1, value="Table")
                sheet.cell(row=row, column=2, value=snapshot.get("tableName", "TABLE 814.3"))
                row += 2
                cond_zones = snapshot.get("zones", [])
                if not isinstance(cond_zones, list):
                    cond_zones = []
                if not cond_zones:
                    legacy_rows = snapshot.get("rows", [])
                    if not isinstance(legacy_rows, list):
                        legacy_rows = []
                    cond_zones = [{"name": "Zone 1", "rows": legacy_rows, "zoneTotalTons": snapshot.get("totalTonsUsed", 0), "recommendedSize": snapshot.get("recommendedSize", ""), "warning": snapshot.get("warning", "")}]

                zone_count = 0
                for zone in cond_zones:
                    if not isinstance(zone, dict):
                        continue
                    zone_count += 1
                    sheet.cell(row=row, column=1, value="Zone")
                    sheet.cell(row=row, column=2, value=zone.get("name", f"Zone {zone_count}"))
                    row += 1
                    headers = ["Equipment", "Tons/Unit", "Quantity", "Row Total Tons", "Row Min Size"]
                    for i, title in enumerate(headers, start=1):
                        sheet.cell(row=row, column=i, value=title)
                    row += 1
                    zone_rows = zone.get("rows", [])
                    if not isinstance(zone_rows, list):
                        zone_rows = []
                    for cond_row in zone_rows:
                        if not isinstance(cond_row, dict):
                            continue
                        sheet.cell(row=row, column=1, value=cond_row.get("equipmentLabel", cond_row.get("equipment", "")))
                        sheet.cell(row=row, column=2, value=cond_row.get("tonsPerUnit", 0))
                        sheet.cell(row=row, column=3, value=cond_row.get("quantity", 0))
                        sheet.cell(row=row, column=4, value=cond_row.get("rowTotalTons", 0))
                        sheet.cell(row=row, column=5, value=cond_row.get("rowSize", ""))
                        row += 1
                    sheet.cell(row=row, column=1, value="Zone Total Tons")
                    sheet.cell(row=row, column=2, value=zone.get("zoneTotalTons", 0))
                    row += 1
                    sheet.cell(row=row, column=1, value="Zone Recommended Size")
                    sheet.cell(row=row, column=2, value=zone.get("recommendedSize", ""))
                    row += 1
                    sheet.cell(row=row, column=1, value="Zone Warning")
                    sheet.cell(row=row, column=2, value=zone.get("warning", ""))
                    row += 2
                data_rows = max(1, zone_count)
            elif section == "vent":
                vent_zones = snapshot.get("zones", [])
                if not isinstance(vent_zones, list):
                    vent_zones = []
                sheet.cell(row=row, column=1, value="Code Basis")
                sheet.cell(row=row, column=2, value=snapshot.get("codeBasis", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Usage Type")
                sheet.cell(row=row, column=2, value=self._fixture_category_label(snapshot.get("usageType", "PRIVATE")))
                row += 1
                sheet.cell(row=row, column=1, value="Drainage Orientation")
                sheet.cell(
                    row=row,
                    column=2,
                    value=("Drainage Vertical" if str(snapshot.get("drainageOrientation", "HORIZONTAL")) == "VERTICAL" else "Drainage Horizontal"),
                )
                row += 1
                sheet.cell(row=row, column=1, value="Effective Total DFU")
                sheet.cell(row=row, column=2, value=snapshot.get("dfu", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Auto Total DFU")
                sheet.cell(row=row, column=2, value=snapshot.get("autoTotalDfu", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Manual Total DFU")
                sheet.cell(row=row, column=2, value=snapshot.get("manualTotalDfu", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Use Manual Override")
                sheet.cell(row=row, column=2, value=bool(snapshot.get("useManualTotal", False)))
                row += 1
                sheet.cell(row=row, column=1, value="Recommended Sanitary Drainage Size")
                sheet.cell(row=row, column=2, value=snapshot.get("recommendedSize", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Sizing Table")
                sheet.cell(row=row, column=2, value=snapshot.get("sizingTable", "TABLE 703.2"))
                row += 1
                sheet.cell(row=row, column=1, value="Table Maximum Units (Selected Size)")
                sheet.cell(row=row, column=2, value=snapshot.get("tableMaxUnits", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Maximum Length Reference")
                sheet.cell(row=row, column=2, value=snapshot.get("maxLengthReference", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Warning")
                sheet.cell(row=row, column=2, value=snapshot.get("warning", ""))
                row += 2
                if vent_zones:
                    zone_count = 0
                    for zone in vent_zones:
                        if not isinstance(zone, dict):
                            continue
                        zone_count += 1
                        sheet.cell(row=row, column=1, value="Zone")
                        sheet.cell(row=row, column=2, value=zone.get("name", f"Zone {zone_count}"))
                        row += 1
                        headers = ["Fixture", "Quantity", "Unit DFU", "Row DFU"]
                        for i, title in enumerate(headers, start=1):
                            sheet.cell(row=row, column=i, value=title)
                        row += 1
                        for vent_row in zone.get("rows", zone.get("fixtures", [])):
                            if not isinstance(vent_row, dict):
                                continue
                            sheet.cell(row=row, column=1, value=vent_row.get("fixtureLabel", vent_row.get("fixtureKey", "")))
                            sheet.cell(row=row, column=2, value=vent_row.get("quantity", 0))
                            sheet.cell(row=row, column=3, value=vent_row.get("unitDfu", 0))
                            sheet.cell(row=row, column=4, value=vent_row.get("rowTotalDfu", 0))
                            row += 1
                        sheet.cell(row=row, column=1, value="Zone DFU")
                        sheet.cell(row=row, column=2, value=zone.get("zoneDfu", 0))
                        row += 2
                    data_rows = max(1, zone_count)
                else:
                    vent_rows = snapshot.get("rows", [])
                    if not isinstance(vent_rows, list):
                        vent_rows = []
                    headers = ["Fixture", "Quantity", "Unit DFU", "Row DFU"]
                    for i, title in enumerate(headers, start=1):
                        sheet.cell(row=row, column=i, value=title)
                    row += 1
                    for vent_row in vent_rows:
                        if not isinstance(vent_row, dict):
                            continue
                        sheet.cell(row=row, column=1, value=vent_row.get("fixtureLabel", vent_row.get("fixtureKey", "")))
                        sheet.cell(row=row, column=2, value=vent_row.get("quantity", 0))
                        sheet.cell(row=row, column=3, value=vent_row.get("unitDfu", 0))
                        sheet.cell(row=row, column=4, value=vent_row.get("rowTotalDfu", 0))
                        row += 1
                    data_rows = max(1, len(vent_rows))
            elif section == "ventilation":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                zones = result.get("zones", selection.get("zones", []))
                if not isinstance(zones, list):
                    zones = []

                workbook = Workbook()
                summary_sheet = workbook.active
                summary_sheet.title = "Ventilation Summary"
                zone_sheet = workbook.create_sheet("Zone Raw Data")
                report_sheet = workbook.create_sheet("Formatted Ventilation Report")
                project_info = self._extract_export_project(snapshot)
                if not isinstance(project_info, dict):
                    project_info = {}
                hdr_fill = PatternFill(fill_type="solid", fgColor="E7EFFA")
                title_fill = PatternFill(fill_type="solid", fgColor="D8E5F6")
                thin = Side(style="thin", color="D1D9E6")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                center = Alignment(horizontal="center", vertical="center")
                summary_value_fill = PatternFill(fill_type="solid", fgColor="EFF6FF")
                final_fill = PatternFill(fill_type="solid", fgColor="E9F7EE")

                system_type = str(selection.get("systemType", result.get("systemType", "SINGLE_ZONE")) or "SINGLE_ZONE")
                if system_type == "ALL_OA":
                    system_type_label = "100% Outdoor Air System"
                elif system_type == "MULTI_RECIRC":
                    system_type_label = "Multiple-zone Recirculating System"
                else:
                    system_type_label = "Single-zone System"

                summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
                title_cell = summary_sheet.cell(row=1, column=1, value="Ventilation Summary")
                title_cell.font = Font(bold=True, size=13)
                title_cell.fill = title_fill
                title_cell.alignment = center

                zone_rows = []
                for zidx, zone in enumerate(zones):
                    if not isinstance(zone, dict):
                        continue
                    occupancy_label = str(zone.get("occupancyCategory", zone.get("occupancyKey", "")) or "")
                    area_val = self._to_non_negative_number(zone.get("area", 0.0), 0.0)
                    pop_mode = "MANUAL" if str(zone.get("populationMode", "DEFAULT")) == "MANUAL" else "DEFAULT"
                    pz_val = self._to_non_negative_number(zone.get("pz", zone.get("manualPopulation", 0.0)), 0.0)
                    rp_val = self._to_non_negative_number(zone.get("rp", 0.0), 0.0)
                    ra_val = self._to_non_negative_number(zone.get("ra", 0.0), 0.0)
                    density_val = self._to_non_negative_number(zone.get("density", 0.0), 0.0)
                    ez_val = self._to_non_negative_number(zone.get("ez", 1.0), 1.0)
                    vbz_val = self._to_non_negative_number(zone.get("vbz", 0.0), 0.0)
                    voz_val = self._to_non_negative_number(zone.get("voz", 0.0), 0.0)
                    vpz_min_val = self._to_non_negative_number(zone.get("vpzMin", 0.0), 0.0)
                    provided_spec = self._as_bool(zone.get("providedSupplySpecified", False), False)
                    provided_val = self._to_non_negative_number(zone.get("providedSupply", 0.0), 0.0)
                    status_val = "—"
                    if provided_spec:
                        status_val = "OK" if provided_val >= voz_val else "Under"
                    zone_rows.append(
                        {
                            "id": str(zone.get("zoneId", zone.get("id", f"zone_{zidx+1}")) or f"zone_{zidx+1}"),
                            "zoneName": str(zone.get("zoneName", zone.get("name", f"Zone {zidx+1}")) or f"Zone {zidx+1}"),
                            "spaceName": str(zone.get("spaceName", "") or ""),
                            "occupancy": occupancy_label,
                            "area": area_val,
                            "populationMode": pop_mode,
                            "populationBasisLabel": ("Manual" if pop_mode == "MANUAL" else "Default"),
                            "pz": pz_val,
                            "rp": rp_val,
                            "ra": ra_val,
                            "density": density_val,
                            "airClass": str(zone.get("airClass", "") or ""),
                            "ezLabel": str(zone.get("ezLabel", "") or ""),
                            "ez": ez_val,
                            "vbz": vbz_val,
                            "voz": voz_val,
                            "vpzMin": vpz_min_val,
                            "providedSupply": provided_val,
                            "providedSupplySpecified": provided_spec,
                            "status": status_val,
                            "notes": str(zone.get("notes", "") or ""),
                        }
                    )

                sum_result = result.get("summary", {})
                if not isinstance(sum_result, dict):
                    sum_result = {}
                total_area = self._to_non_negative_number(sum_result.get("totalArea", sum_result.get("sumArea", 0.0)), 0.0) if sum_result else 0.0
                total_pz = self._to_non_negative_number(sum_result.get("sumPz", result.get("sumPz", 0.0)), 0.0)
                total_rppz = self._to_non_negative_number(sum_result.get("sumRpPz", 0.0), 0.0)
                total_raaz = self._to_non_negative_number(sum_result.get("sumRaAz", 0.0), 0.0)
                total_vbz = self._to_non_negative_number(sum_result.get("sumVbz", 0.0), 0.0)
                total_voz = self._to_non_negative_number(sum_result.get("sumVoz", result.get("sumVoz", 0.0)), 0.0)
                if not sum_result and zone_rows:
                    total_area = sum(row["area"] for row in zone_rows)
                    total_pz = sum(row["pz"] for row in zone_rows)
                    total_rppz = sum((row["rp"] * row["pz"]) for row in zone_rows)
                    total_raaz = sum((row["ra"] * row["area"]) for row in zone_rows)
                    total_vbz = sum(row["vbz"] for row in zone_rows)
                    total_voz = sum(row["voz"] for row in zone_rows)
                summary_rows = [
                    ("Project", project_info.get("projectName", selection.get("projectName", ""))),
                    ("Project Number", project_info.get("projectNumber", "")),
                    ("Client", project_info.get("clientName", "")),
                    ("Location", project_info.get("projectLocation", "")),
                    ("System Type", system_type_label),
                    ("Status", result.get("status", "incomplete")),
                    ("Final Intake Vot (CFM)", result.get("Vot", 0)),
                    ("Final Intake Vot (L/s)", self._to_non_negative_number(result.get("Vot", 0), 0.0) * 0.47194745),
                    ("Uncorrected Outdoor Airflow Vou (CFM)", result.get("Vou", 0)),
                    ("System Ventilation Efficiency Ev", result.get("Ev", 0)),
                    ("Occupant Diversity D", result.get("D", 0)),
                    ("System Population Ps", selection.get("systemPopulationPs", result.get("ps", 0))),
                    ("Total Area ΣAz (ft²)", total_area),
                    ("Total Zone Population ΣPz", total_pz),
                    ("Σ(Rp × Pz)", total_rppz),
                    ("Σ(Ra × Az)", total_raaz),
                    ("ΣVbz (CFM)", total_vbz),
                    ("ΣVoz (CFM)", total_voz),
                    ("Exported At", now_text),
                    ("App Version", str(snapshot.get("app_version", self._APP_VERSION) or self._APP_VERSION)),
                    ("Notes", "Verify code applicability, jurisdictional amendments, and energy-code requirements before final design use."),
                ]
                sr = 3
                final_rows = {"Final Intake Vot (CFM)", "Final Intake Vot (L/s)"}
                for key, value in summary_rows:
                    key_cell = summary_sheet.cell(row=sr, column=1, value=key)
                    val_cell = summary_sheet.cell(row=sr, column=2, value=value)
                    key_cell.font = Font(bold=True)
                    key_cell.border = border
                    val_cell.border = border
                    key_cell.alignment = Alignment(horizontal="left", vertical="center")
                    val_cell.alignment = Alignment(horizontal="left", vertical="center")
                    val_cell.fill = final_fill if key in final_rows else summary_value_fill
                    sr += 1
                summary_sheet.freeze_panes = "A3"
                summary_sheet.column_dimensions["A"].width = 40
                summary_sheet.column_dimensions["B"].width = 28

                z_headers = [
                    "Zone ID",
                    "Zone name",
                    "Space name",
                    "Occupancy category",
                    "Area (ft²)",
                    "Population basis",
                    "Population Pz",
                    "Rp",
                    "Ra",
                    "Density (people/1000 ft²)",
                    "Air class",
                    "Ez label",
                    "Ez",
                    "Vbz (CFM)",
                    "Voz (CFM)",
                    "Vpz-min (CFM)",
                    "Provided Supply (CFM)",
                    "Status",
                    "Notes",
                ]
                for i, h in enumerate(z_headers, start=1):
                    cell = zone_sheet.cell(row=1, column=i, value=h)
                    cell.font = Font(bold=True)
                    cell.fill = hdr_fill
                    cell.alignment = center
                    cell.border = border
                zr = 2
                for zone in zone_rows:
                    values = [
                        zone["id"],
                        zone["zoneName"],
                        zone["spaceName"],
                        zone["occupancy"],
                        zone["area"],
                        zone["populationBasisLabel"],
                        zone["pz"],
                        zone["rp"],
                        zone["ra"],
                        zone["density"],
                        zone["airClass"],
                        zone["ezLabel"],
                        zone["ez"],
                        zone["vbz"],
                        zone["voz"],
                        zone["vpzMin"],
                        zone["providedSupply"] if zone["providedSupplySpecified"] else "",
                        zone["status"],
                        zone["notes"],
                    ]
                    for c, v in enumerate(values, start=1):
                        cell = zone_sheet.cell(row=zr, column=c, value=v)
                        cell.border = border
                        if c >= 5 and c <= 18:
                            cell.alignment = center
                        if c == 18 and str(v).upper() == "OK":
                            cell.fill = PatternFill(fill_type="solid", fgColor="EAF6EE")
                        elif c == 18 and str(v).upper() == "UNDER":
                            cell.fill = PatternFill(fill_type="solid", fgColor="FDECEC")
                    zr += 1
                zone_sheet.freeze_panes = "A2"
                for c in range(1, len(z_headers) + 1):
                    letter = zone_sheet.cell(row=1, column=c).column_letter
                    if c == 19:
                        zone_sheet.column_dimensions[letter].width = 30
                    elif c in (2, 3, 4, 12):
                        zone_sheet.column_dimensions[letter].width = 22
                    elif c in (18,):
                        zone_sheet.column_dimensions[letter].width = 12
                    else:
                        zone_sheet.column_dimensions[letter].width = 14
                for rr in range(2, zr):
                    for cc in [5, 7, 8, 9, 10, 13, 14, 15, 16, 17]:
                        zone_sheet.cell(row=rr, column=cc).number_format = "0.00"

                # Formatted engineering report worksheet.
                report_headers = [
                    "ROOM NAME",
                    "FLOOR AREA (ft²)",
                    "OUTDOOR AIRFLOW RATE (cfm/ft²)",
                    "PEOPLE OUTDOOR RATE (cfm/person)",
                    "DEFAULT DENSITY (/1000 ft²)",
                    "POPULATION BASIS",
                    "NO. OF PERSONS Pz",
                    "REQUIRED VENTILATION Vbz (cfm)",
                    "ZONE DISTRIBUTION EFFECTIVENESS Ez",
                    "TOTAL REQUIRED VENTILATION Voz (cfm)",
                    "PROVIDED SUPPLY (cfm)",
                ]
                report_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report_headers))
                report_title = report_sheet.cell(row=1, column=1, value="COMMON AREAS VENTILATION (CMC TABLE 402.1)")
                report_title.font = Font(bold=True, size=13)
                report_title.alignment = center
                report_title.fill = title_fill
                report_sheet.row_dimensions[1].height = 24

                header_row = 3
                for i, header in enumerate(report_headers, start=1):
                    hcell = report_sheet.cell(row=header_row, column=i, value=header)
                    hcell.font = Font(bold=True)
                    hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    hcell.fill = hdr_fill
                    hcell.border = border
                report_sheet.row_dimensions[header_row].height = 34

                report_row = header_row + 1
                for row_data in zone_rows:
                    values = [
                        row_data["zoneName"],
                        row_data["area"],
                        row_data["ra"],
                        row_data["rp"],
                        row_data["density"] if row_data["populationMode"] != "MANUAL" else "",
                        row_data["populationBasisLabel"],
                        row_data["pz"],
                        row_data["vbz"],
                        row_data["ez"],
                        row_data["voz"],
                        row_data["providedSupply"] if row_data["providedSupplySpecified"] else "",
                    ]
                    for col_idx, value in enumerate(values, start=1):
                        cell = report_sheet.cell(row=report_row, column=col_idx, value=value)
                        cell.border = border
                        if col_idx > 1:
                            cell.alignment = center
                    report_row += 1

                totals_row = report_row + 1
                report_sheet.cell(row=totals_row, column=1, value="TOTAL AREA (ft²)").font = Font(bold=True)
                report_sheet.cell(row=totals_row, column=2, value=total_area).font = Font(bold=True)
                report_sheet.cell(row=totals_row, column=2).number_format = "0.00"
                report_sheet.cell(row=totals_row, column=6, value="TOTAL POPULATION").font = Font(bold=True)
                report_sheet.cell(row=totals_row, column=7, value=total_pz).font = Font(bold=True)
                report_sheet.cell(row=totals_row, column=7).number_format = "0.00"
                report_sheet.cell(row=totals_row + 1, column=1, value="SUM Vbz (cfm)").font = Font(bold=True)
                report_sheet.cell(row=totals_row + 1, column=2, value=total_vbz).font = Font(bold=True)
                report_sheet.cell(row=totals_row + 1, column=2).number_format = "0.00"
                report_sheet.cell(row=totals_row + 1, column=6, value="SUM Voz (cfm)").font = Font(bold=True)
                report_sheet.cell(row=totals_row + 1, column=7, value=total_voz).font = Font(bold=True)
                report_sheet.cell(row=totals_row + 1, column=7).number_format = "0.00"

                report_sheet.freeze_panes = "A4"
                report_widths = [28, 16, 20, 20, 19, 17, 14, 22, 22, 22, 18]
                for i, width in enumerate(report_widths, start=1):
                    report_sheet.column_dimensions[report_sheet.cell(row=header_row, column=i).column_letter].width = width

                for rr in range(header_row + 1, report_row):
                    for cc in [2, 3, 4, 5, 7, 8, 9, 10, 11]:
                        report_sheet.cell(row=rr, column=cc).number_format = "0.00"

                data_rows = max(1, len(zone_rows))
            elif section == "gas":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                lines = result.get("lines", [])
                if not isinstance(lines, list):
                    lines = []

                sheet.cell(row=row, column=1, value="Code Basis")
                sheet.cell(row=row, column=2, value=selection.get("codeBasis", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Gas Type")
                sheet.cell(row=row, column=2, value=selection.get("gasType", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Material")
                sheet.cell(row=row, column=2, value=selection.get("material", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Inlet Pressure (psi)")
                sheet.cell(row=row, column=2, value=selection.get("inletPressure", 0))
                row += 1
                if str(selection.get("inletPressureOption", "") or "").upper() == "LT2":
                    sheet.cell(row=row, column=1, value="Inlet Pressure Option")
                    sheet.cell(row=row, column=2, value="Less Than 2")
                    row += 1
                sheet.cell(row=row, column=1, value="Pressure Drop")
                sheet.cell(row=row, column=2, value=selection.get("pressureDrop", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Specific Gravity")
                sheet.cell(row=row, column=2, value=selection.get("specificGravity", 0))
                row += 2

                headers = [
                    "Line",
                    "Demand (CFH)",
                    "Run Length (ft)",
                    "Recommended Size",
                    "Capacity (CFH)",
                    "Max Allowable Length (ft)",
                    "Warning",
                ]
                for i, title in enumerate(headers, start=1):
                    sheet.cell(row=row, column=i, value=title)
                row += 1
                for line in lines:
                    if not isinstance(line, dict):
                        continue
                    sheet.cell(row=row, column=1, value=line.get("lineLabel", ""))
                    sheet.cell(row=row, column=2, value=line.get("demandCfh", 0))
                    sheet.cell(row=row, column=3, value=line.get("runLength", 0))
                    sheet.cell(row=row, column=4, value=line.get("recommendedSize", ""))
                    sheet.cell(row=row, column=5, value=line.get("capacityCfh", 0))
                    sheet.cell(row=row, column=6, value=line.get("maxAllowableLength", 0))
                    sheet.cell(row=row, column=7, value=line.get("warning", ""))
                    row += 1
                data_rows = len(lines)
            elif section == "wsfu":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                wsfu_zones = result.get("zones", [])
                if not isinstance(wsfu_zones, list):
                    wsfu_zones = []
                flush_type = "FLUSH_VALVE" if str(result.get("flushType", "FLUSH_TANK")) == "FLUSH_VALVE" else "FLUSH_TANK"
                sizing_basis = "Flush Valve" if flush_type == "FLUSH_VALVE" else "Flush Tank"
                selected_total_gpm = (
                    self._to_non_negative_number(result.get("flushValveGpm", 0), 0.0)
                    if flush_type == "FLUSH_VALVE"
                    else self._to_non_negative_number(result.get("flushTankGpm", 0), 0.0)
                )
                selected_matched_fu = (
                    self._to_non_negative_number(result.get("matchedValveFu", 0), 0.0)
                    if flush_type == "FLUSH_VALVE"
                    else self._to_non_negative_number(result.get("matchedTankFu", 0), 0.0)
                )
                sheet.cell(row=row, column=1, value="Total Fixture Units")
                sheet.cell(row=row, column=2, value=result.get("totalFu", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Code Basis")
                sheet.cell(row=row, column=2, value=selection.get("codeBasis", "IPC"))
                row += 1
                sheet.cell(row=row, column=1, value="Usage Type")
                sheet.cell(row=row, column=2, value=self._fixture_category_label(selection.get("usageType", "PRIVATE")))
                row += 1
                sheet.cell(row=row, column=1, value="Use Manual Total")
                sheet.cell(row=row, column=2, value=self._as_bool(selection.get("useManualTotal", False), False))
                row += 1
                sheet.cell(row=row, column=1, value="Auto Total WSFU")
                sheet.cell(row=row, column=2, value=self._to_non_negative_number(selection.get("autoTotalFu", 0), 0))
                row += 1
                sheet.cell(row=row, column=1, value="Manual Total WSFU")
                sheet.cell(row=row, column=2, value=self._to_non_negative_number(selection.get("manualTotalFu", 0), 0))
                row += 1
                sheet.cell(row=row, column=1, value="Sizing Basis")
                sheet.cell(row=row, column=2, value=sizing_basis)
                row += 1
                sheet.cell(row=row, column=1, value="Selected Flow (GPM)")
                sheet.cell(row=row, column=2, value=result.get("selectedFlowGpm", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Design Length (ft)")
                sheet.cell(row=row, column=2, value=result.get("designLengthFt", 100))
                row += 1
                sheet.cell(row=row, column=1, value=f"Total GPM ({sizing_basis})")
                sheet.cell(row=row, column=2, value=selected_total_gpm)
                row += 1
                sheet.cell(row=row, column=1, value=f"Matched FU ({sizing_basis})")
                sheet.cell(row=row, column=2, value=selected_matched_fu)
                row += 1
                sheet.cell(row=row, column=1, value="Water Service Size")
                sheet.cell(row=row, column=2, value=result.get("serviceSize", "Manual selection required"))
                row += 1
                sheet.cell(row=row, column=1, value="Size Capacity (GPM)")
                sheet.cell(row=row, column=2, value=result.get("sizeCapacityGpm", 0))
                row += 1
                warning = str(result.get("warning", "") or "")
                if warning:
                    sheet.cell(row=row, column=1, value="Warning")
                    sheet.cell(row=row, column=2, value=warning)
                    row += 1
                sheet.cell(row=row, column=1, value="Source")
                sheet.cell(row=row, column=2, value=result.get("source", ""))
                row += 2
                fixture_rows = selection.get("fixtureRows", [])
                if not isinstance(fixture_rows, list):
                    fixture_rows = []
                fixture_headers = ["Fixture", "Quantity", "Hot+Cold (0.75)", "Unit WSFU", "Row Total WSFU"]
                if wsfu_zones:
                    zone_rows_count = 0
                    for zidx, zone in enumerate(wsfu_zones):
                        if not isinstance(zone, dict):
                            continue
                        sheet.cell(row=row, column=1, value="Zone")
                        sheet.cell(row=row, column=2, value=zone.get("name", f"Zone {zidx+1}"))
                        row += 1
                        for i, title in enumerate(fixture_headers, start=1):
                            sheet.cell(row=row, column=i, value=title)
                        row += 1
                        for f_row in zone.get("fixtureRows", []):
                            if not isinstance(f_row, dict):
                                continue
                            zone_rows_count += 1
                            sheet.cell(row=row, column=1, value=f_row.get("fixtureLabel", ""))
                            sheet.cell(row=row, column=2, value=f_row.get("quantity", 0))
                            sheet.cell(row=row, column=3, value=bool(f_row.get("hotCold", False)))
                            sheet.cell(row=row, column=4, value=f_row.get("unitWsfu", 0))
                            sheet.cell(row=row, column=5, value=f_row.get("rowTotalWsfu", 0))
                            row += 1
                        sheet.cell(row=row, column=1, value="Zone Subtotal WSFU")
                        sheet.cell(row=row, column=2, value=zone.get("totalFu", 0))
                        row += 2
                    if zone_rows_count <= 0:
                        row += 1
                else:
                    for i, title in enumerate(fixture_headers, start=1):
                        sheet.cell(row=row, column=i, value=title)
                    row += 1
                    for f_row in fixture_rows:
                        if not isinstance(f_row, dict):
                            continue
                        sheet.cell(row=row, column=1, value=f_row.get("fixtureLabel", ""))
                        sheet.cell(row=row, column=2, value=f_row.get("quantity", 0))
                        sheet.cell(row=row, column=3, value=bool(f_row.get("hotCold", False)))
                        sheet.cell(row=row, column=4, value=f_row.get("unitWsfu", 0))
                        sheet.cell(row=row, column=5, value=f_row.get("rowTotalWsfu", 0))
                        row += 1
                    if fixture_rows:
                        row += 1

                friction_rows = result.get("frictionLossByPipeType", [])
                if not isinstance(friction_rows, list):
                    friction_rows = []
                headers = [
                    "Pipe Type",
                    "Hazen-Williams C",
                    "Head Loss (ft/100ft)",
                    "Pressure Loss (psi/100ft)",
                    "Pressure Loss at Design Length (psi)",
                ]
                for i, title in enumerate(headers, start=1):
                    sheet.cell(row=row, column=i, value=title)
                row += 1
                for f_row in friction_rows:
                    if not isinstance(f_row, dict):
                        continue
                    sheet.cell(row=row, column=1, value=f_row.get("pipeType", ""))
                    sheet.cell(row=row, column=2, value=f_row.get("cFactor", 0))
                    sheet.cell(row=row, column=3, value=f_row.get("headLossFtPer100", 0))
                    sheet.cell(row=row, column=4, value=f_row.get("psiLossPer100", 0))
                    sheet.cell(row=row, column=5, value=f_row.get("psiLossForLength", 0))
                    row += 1
                data_rows = max(1, len(friction_rows), len(fixture_rows))
            elif section == "fixtureunit":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}

                sanitary = result.get("sanitary", {})
                if not isinstance(sanitary, dict):
                    sanitary = {}
                water = result.get("water", {})
                if not isinstance(water, dict):
                    water = {}
                zones = selection.get("zones", [])
                if not isinstance(zones, list):
                    zones = []
                rows = selection.get("rows", [])
                if not isinstance(rows, list):
                    rows = []
                if not zones:
                    zones = [{"name": "Zone 1", "rows": rows}]

                sheet.cell(row=row, column=1, value="Code Basis")
                sheet.cell(row=row, column=2, value=selection.get("codeBasis", "IPC"))
                row += 1
                sheet.cell(row=row, column=1, value="Drainage Orientation")
                sheet.cell(row=row, column=2, value=("Vertical" if str(selection.get("drainageOrientation", "HORIZONTAL")) == "VERTICAL" else "Horizontal"))
                row += 1
                sheet.cell(row=row, column=1, value="Water Sizing Basis")
                sheet.cell(row=row, column=2, value=("Flush Valve" if str(selection.get("flushType", "FLUSH_TANK")) == "FLUSH_VALVE" else "Flush Tank"))
                row += 1
                sheet.cell(row=row, column=1, value="Water Design Length (ft)")
                sheet.cell(row=row, column=2, value=selection.get("designLengthFt", 100))
                row += 2

                zone_rows_count = 0
                headers = [
                    "Fixture",
                    "Category",
                    "Quantity",
                    "Hot+Cold",
                    "Note",
                    "Waste / Fixture",
                    "Water / Fixture",
                    "Row Waste",
                    "Row Water",
                ]
                for zone_index, zone in enumerate(zones):
                    if not isinstance(zone, dict):
                        continue
                    zone_name = str(zone.get("name", f"Zone {zone_index+1}") or f"Zone {zone_index+1}")
                    z_rows = zone.get("rows", [])
                    if not isinstance(z_rows, list):
                        z_rows = []
                    sheet.cell(row=row, column=1, value="Zone")
                    sheet.cell(row=row, column=2, value=zone_name)
                    row += 1
                    for i, title in enumerate(headers, start=1):
                        sheet.cell(row=row, column=i, value=title)
                    row += 1
                    zone_waste = 0.0
                    zone_water = 0.0
                    for item in z_rows:
                        if not isinstance(item, dict):
                            continue
                        zone_rows_count += 1
                        waste_total = self._to_non_negative_number(item.get("rowWasteTotal", item.get("rowSanitaryTotal", 0)), 0.0)
                        water_total = self._to_non_negative_number(item.get("rowWaterTotal", 0), 0.0)
                        zone_waste += waste_total
                        zone_water += water_total
                        sheet.cell(row=row, column=1, value=item.get("fixtureLabel", item.get("fixtureKey", "")))
                        sheet.cell(row=row, column=2, value=self._fixture_category_label(item.get("category", "PRIVATE")))
                        sheet.cell(row=row, column=3, value=item.get("quantity", 0))
                        sheet.cell(row=row, column=4, value=bool(item.get("hotCold", False)))
                        sheet.cell(row=row, column=5, value=item.get("note", ""))
                        sheet.cell(row=row, column=6, value=item.get("wasteUnit", item.get("sanitaryUnit", 0)))
                        sheet.cell(row=row, column=7, value=item.get("waterUnit", 0))
                        sheet.cell(row=row, column=8, value=waste_total)
                        sheet.cell(row=row, column=9, value=water_total)
                        row += 1
                    sheet.cell(row=row, column=1, value="Zone Waste Total")
                    sheet.cell(row=row, column=2, value=zone_waste)
                    row += 1
                    sheet.cell(row=row, column=1, value="Zone Water Total")
                    sheet.cell(row=row, column=2, value=zone_water)
                    row += 2

                sheet.cell(row=row, column=1, value="Sanitary Summary")
                row += 1
                sheet.cell(row=row, column=1, value="Total Sanitary Load")
                sheet.cell(row=row, column=2, value=sanitary.get("total", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Suggested Pipe Size")
                sheet.cell(row=row, column=2, value=sanitary.get("recommendedSize", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Sanitary Warning")
                sheet.cell(row=row, column=2, value=sanitary.get("warning", ""))
                row += 2

                sheet.cell(row=row, column=1, value="Water Supply Summary")
                row += 1
                sheet.cell(row=row, column=1, value="Total WSFU")
                sheet.cell(row=row, column=2, value=water.get("total", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Estimated Demand (GPM)")
                sheet.cell(row=row, column=2, value=water.get("selectedFlowGpm", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Suggested Pipe Size")
                sheet.cell(row=row, column=2, value=water.get("serviceSize", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Water Warning")
                sheet.cell(row=row, column=2, value=water.get("warning", ""))
                data_rows = max(1, zone_rows_count)
            elif section == "duct":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                duct_zones = result.get("zones", [])
                if not isinstance(duct_zones, list):
                    duct_zones = []
                sheet.cell(row=row, column=1, value="Mode")
                sheet.cell(row=row, column=2, value=selection.get("mode", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Shape")
                sheet.cell(row=row, column=2, value=selection.get("shape", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Airflow (CFM)")
                sheet.cell(row=row, column=2, value=selection.get("cfm", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Friction Target (in.wg/100ft)")
                sheet.cell(row=row, column=2, value=selection.get("frictionTarget", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Round Diameter (in)")
                sheet.cell(row=row, column=2, value=selection.get("diameter", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Rect Width (in)")
                sheet.cell(row=row, column=2, value=selection.get("width", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Rect Height (in)")
                sheet.cell(row=row, column=2, value=selection.get("height", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Aspect Ratio Limit")
                sheet.cell(row=row, column=2, value=selection.get("ratioLimit", 0))
                row += 2
                if duct_zones:
                    headers = ["Zone", "Segment", "Mode", "Eq Diameter (in)", "Round", "Rect", "Velocity (FPM)", "Friction (in.wg/100ft)", "Warning"]
                    for i, title in enumerate(headers, start=1):
                        sheet.cell(row=row, column=i, value=title)
                    row += 1
                    line_count = 0
                    for zone in duct_zones:
                        if not isinstance(zone, dict):
                            continue
                        zone_name = zone.get("name", "")
                        for r in zone.get("rows", []):
                            if not isinstance(r, dict):
                                continue
                            line_count += 1
                            sheet.cell(row=row, column=1, value=zone_name)
                            sheet.cell(row=row, column=2, value=r.get("name", ""))
                            sheet.cell(row=row, column=3, value=r.get("modeLabel", ""))
                            sheet.cell(row=row, column=4, value=r.get("equivalentDiameter", 0))
                            sheet.cell(row=row, column=5, value=r.get("recommendedRound", ""))
                            sheet.cell(row=row, column=6, value=r.get("recommendedRect", ""))
                            sheet.cell(row=row, column=7, value=r.get("velocityFpm", 0))
                            sheet.cell(row=row, column=8, value=r.get("frictionRate", 0))
                            sheet.cell(row=row, column=9, value=r.get("warning", ""))
                            row += 1
                    data_rows = max(1, line_count)
                else:
                    sheet.cell(row=row, column=1, value="Status")
                    sheet.cell(row=row, column=2, value=result.get("status", ""))
                    row += 1
                    sheet.cell(row=row, column=1, value="Message")
                    sheet.cell(row=row, column=2, value=result.get("message", ""))
                    row += 1
                    sheet.cell(row=row, column=1, value="Equivalent Diameter (in)")
                    sheet.cell(row=row, column=2, value=result.get("equivalentDiameter", 0))
                    row += 1
                    sheet.cell(row=row, column=1, value="Recommended Round")
                    sheet.cell(row=row, column=2, value=result.get("recommendedRound", ""))
                    row += 1
                    sheet.cell(row=row, column=1, value="Recommended Rect")
                    sheet.cell(row=row, column=2, value=result.get("recommendedRect", ""))
                    row += 1
                    sheet.cell(row=row, column=1, value="Velocity (FPM)")
                    sheet.cell(row=row, column=2, value=result.get("velocityFpm", 0))
                    row += 1
                    sheet.cell(row=row, column=1, value="Velocity Pressure (in.wg)")
                    sheet.cell(row=row, column=2, value=result.get("velocityPressure", 0))
                    row += 1
                    sheet.cell(row=row, column=1, value="Friction Rate (in.wg/100ft)")
                    sheet.cell(row=row, column=2, value=result.get("frictionRate", 0))
                    row += 1
                    sheet.cell(row=row, column=1, value="Warning")
                    sheet.cell(row=row, column=2, value=result.get("warning", ""))
                    data_rows = 1
            elif section == "ductstatic":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                segments = selection.get("segments", [])
                if not isinstance(segments, list):
                    segments = []
                sheet.cell(row=row, column=1, value="Module")
                sheet.cell(row=row, column=2, value="Ducts Static Pressure Calculator")
                row += 2
                headers = [
                    "Segment Name",
                    "Line Type",
                    "Length (ft)",
                    "Friction Rate",
                    "Width",
                    "Height",
                    "Radius",
                    "Velocity (FPM)",
                    "R/W",
                    "H/W",
                    "C Coefficient",
                    "Segment Pressure Drop (in.wg)",
                    "Warning",
                ]
                for i, title in enumerate(headers, start=1):
                    sheet.cell(row=row, column=i, value=title)
                row += 1
                for seg in segments:
                    if not isinstance(seg, dict):
                        continue
                    sheet.cell(row=row, column=1, value=seg.get("name", ""))
                    sheet.cell(row=row, column=2, value=seg.get("lineType", ""))
                    sheet.cell(row=row, column=3, value=seg.get("lengthFt", 0))
                    sheet.cell(row=row, column=4, value=seg.get("frictionRate", 0))
                    sheet.cell(row=row, column=5, value=seg.get("width", 0))
                    sheet.cell(row=row, column=6, value=seg.get("height", 0))
                    sheet.cell(row=row, column=7, value=seg.get("radius", 0))
                    sheet.cell(row=row, column=8, value=seg.get("velocityFpm", 0))
                    sheet.cell(row=row, column=9, value=seg.get("rwRatio", 0))
                    sheet.cell(row=row, column=10, value=seg.get("hwRatio", 0))
                    sheet.cell(row=row, column=11, value=seg.get("cCoefficient", 0))
                    sheet.cell(row=row, column=12, value=seg.get("segmentPressureDrop", 0))
                    sheet.cell(row=row, column=13, value=seg.get("warning", ""))
                    row += 1
                row += 1
                sheet.cell(row=row, column=1, value="Total Pressure Drop (in.wg)")
                sheet.cell(row=row, column=2, value=result.get("totalPressureDrop", 0))
                warnings = result.get("warnings", [])
                if not isinstance(warnings, list):
                    warnings = []
                if warnings:
                    row += 1
                    sheet.cell(row=row, column=1, value="Warnings")
                    sheet.cell(row=row, column=2, value=" | ".join([str(w) for w in warnings if str(w).strip()]))
                data_rows = max(1, len(segments))
            elif section == "solar":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                sheet.cell(row=row, column=1, value="Building SF (CFA)")
                sheet.cell(row=row, column=2, value=selection.get("buildingSf", 0))
                row += 1
                sheet.cell(row=row, column=1, value="ZIP")
                sheet.cell(row=row, column=2, value=selection.get("zipCode", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Climate Zone")
                sheet.cell(row=row, column=2, value=selection.get("climateZone", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Building Type")
                sheet.cell(row=row, column=2, value=selection.get("buildingType", ""))
                row += 1
                sheet.cell(row=row, column=1, value="SARA Input")
                sheet.cell(row=row, column=2, value=selection.get("sara", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Use SARA Path")
                sheet.cell(row=row, column=2, value=selection.get("useSaraPath", False))
                row += 1
                sheet.cell(row=row, column=1, value="D Value")
                sheet.cell(row=row, column=2, value=selection.get("dValue", 0.95))
                row += 2

                sheet.cell(row=row, column=1, value="Status")
                sheet.cell(row=row, column=2, value=result.get("status", "pending_workbook"))
                row += 1
                sheet.cell(row=row, column=1, value="Message")
                sheet.cell(row=row, column=2, value=result.get("message", "Solar workbook source is pending."))
                row += 1
                sheet.cell(row=row, column=1, value="Climate Group")
                sheet.cell(row=row, column=2, value=result.get("climateGroup", ""))
                row += 1
                sheet.cell(row=row, column=1, value="Factor A")
                sheet.cell(row=row, column=2, value=result.get("factorA", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Factor B")
                sheet.cell(row=row, column=2, value=result.get("factorB", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Factor C")
                sheet.cell(row=row, column=2, value=result.get("factorC", 0))
                row += 1
                sheet.cell(row=row, column=1, value="PV kW (CFA)")
                sheet.cell(row=row, column=2, value=result.get("pvKwByCfa", 0))
                row += 1
                sheet.cell(row=row, column=1, value="PV kW (SARA)")
                sheet.cell(row=row, column=2, value=result.get("pvKwBySara", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Total Solar kW Required")
                sheet.cell(row=row, column=2, value=result.get("pvKwdc", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Total Battery kWh Required")
                sheet.cell(row=row, column=2, value=result.get("batteryKwh", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Total Battery kW Required")
                sheet.cell(row=row, column=2, value=result.get("batteryKw", 0))
                row += 1
                sheet.cell(row=row, column=1, value="Solar Exemption")
                sheet.cell(row=row, column=2, value=result.get("solarExemptionStatus", "NOT EXEMPT"))
                row += 1
                sheet.cell(row=row, column=1, value="Battery Exemption")
                sheet.cell(row=row, column=2, value=result.get("batteryExemptionStatus", "NOT EXEMPT"))
                row += 1
                sheet.cell(row=row, column=1, value="Notes")
                sheet.cell(row=row, column=2, value=result.get("notes", ""))
                data_rows = 1
            elif section == "refrigerant":
                selection = snapshot.get("selection", {})
                if not isinstance(selection, dict):
                    selection = {}
                result = snapshot.get("result", {})
                if not isinstance(result, dict):
                    result = {}
                classification = selection.get("classification", {})
                if not isinstance(classification, dict):
                    classification = {}
                charge_inputs = selection.get("chargeInputs", {})
                if not isinstance(charge_inputs, dict):
                    charge_inputs = {}
                charge_result = result.get("charge", {})
                if not isinstance(charge_result, dict):
                    charge_result = {}
                ashrae15_inputs = selection.get("ashrae15Inputs", {})
                if not isinstance(ashrae15_inputs, dict):
                    ashrae15_inputs = {}
                ashrae152_inputs = selection.get("ashrae152Inputs", {})
                if not isinstance(ashrae152_inputs, dict):
                    ashrae152_inputs = {}
                advanced_inputs = selection.get("advancedInputs", {})
                if not isinstance(advanced_inputs, dict):
                    advanced_inputs = {}
                details = result.get("details", {})
                if not isinstance(details, dict):
                    details = {}
                multi_zone = selection.get("multiZone", {})
                if not isinstance(multi_zone, dict):
                    multi_zone = {}
                multi_shared = multi_zone.get("shared", {})
                if not isinstance(multi_shared, dict):
                    multi_shared = {}
                multi_rows = result.get("rows", [])
                if not isinstance(multi_rows, list):
                    multi_rows = []

                mode_raw = str(result.get("mode") or selection.get("calculationMode") or "SINGLE").strip().upper()
                is_multi_mode = mode_raw in {"MULTI_ZONE", "MULTI_ZONE_TABLE", "MULTIZONETABLE", "MULTI-ZONE TABLE"} or "MULTI" in mode_raw

                workbook = Workbook()
                sheet = workbook.active
                project_info = self._extract_export_project(snapshot)
                if not isinstance(project_info, dict):
                    project_info = {}

                if is_multi_mode:
                    sheet.title = "A2L Refrigerant Calculation"
                    thin = Side(style="thin", color="B7C7DB")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    header_fill = PatternFill(fill_type="solid", fgColor="E9EFF8")
                    title_fill = PatternFill(fill_type="solid", fgColor="D8E5F6")
                    yes_fill = PatternFill(fill_type="solid", fgColor="E9F7EE")
                    no_fill = PatternFill(fill_type="solid", fgColor="FDECEC")

                    headers = [
                        "Room ID",
                        "Room Volume (ft3)",
                        "Room Area (ft2)",
                        "Space Height (ft)",
                        "Pipe Length (ft)",
                        "Pipe Size",
                        "Number of Units",
                        "Charge per Foot (lb/ft)",
                        "Factory Charge (lb)",
                        "Line Charge (lb)",
                        "Total Charge (lb)",
                        "mREL (lb)",
                        "m1 (lb)",
                        "m2 (lb)",
                        "C",
                        "M",
                        "hc",
                        "Mc (lb)",
                        "Mv (lb)",
                        "mMAX No Vent (lb)",
                        "mMAX With Vent (lb)",
                        "Compliance (No Vent)",
                        "Compliance (With Vent)",
                    ]
                    end_col = len(headers)
                    r = 1
                    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=end_col)
                    t = sheet.cell(row=r, column=1, value="A2L Refrigerant Calculation")
                    t.font = Font(bold=True, size=13)
                    t.alignment = center
                    t.fill = title_fill
                    t.border = border
                    r += 2

                    meta_rows = [
                        ("Project", project_info.get("projectName", "")),
                        ("Refrigerant", multi_shared.get("refrigerantType", classification.get("refrigerantType", ""))),
                        ("Standard", result.get("applicableStandard", "")),
                        ("Data Version", multi_shared.get("dataVersion", classification.get("dataVersion", ""))),
                        ("Exported At", now_text),
                    ]
                    for label, value in meta_rows:
                        sheet.cell(row=r, column=1, value=label).font = Font(bold=True)
                        sheet.cell(row=r, column=2, value=value)
                        r += 1
                    r += 1

                    header_row = r
                    for idx, title in enumerate(headers, start=1):
                        c = sheet.cell(row=header_row, column=idx, value=title)
                        c.font = Font(bold=True)
                        c.alignment = center
                        c.fill = header_fill
                        c.border = border
                    r += 1

                    for item in multi_rows:
                        if not isinstance(item, dict):
                            continue
                        no_val = str(item.get("complianceNoVent", "")).strip().upper()
                        with_val = str(item.get("complianceWithVent", "")).strip().upper()
                        if no_val not in {"YES", "NO"}:
                            s_no = str(item.get("statusNoVent", "")).upper()
                            no_val = "YES" if s_no == "COMPLIANT" else ("NO" if s_no == "NOT COMPLIANT" else "")
                        if with_val not in {"YES", "NO"}:
                            s_with = str(item.get("statusWithVent", "")).upper()
                            with_val = "YES" if s_with == "COMPLIANT" else ("NO" if s_with == "NOT COMPLIANT" else "")

                        values = [
                            item.get("roomId", ""),
                            item.get("roomVolume", 0),
                            item.get("roomArea", 0),
                            item.get("spaceHeight", 0),
                            item.get("pipeLength", 0),
                            item.get("pipeSize", ""),
                            item.get("numberOfUnits", 0),
                            item.get("chargePerFoot", 0),
                            item.get("factoryCharge", 0),
                            item.get("lineCharge", 0),
                            item.get("totalCharge", 0),
                            item.get("mrel", 0),
                            item.get("m1", 0),
                            item.get("m2", 0),
                            item.get("c", 0),
                            item.get("m", 0),
                            item.get("hc", 0),
                            item.get("mc", 0),
                            item.get("mv", 0),
                            item.get("mmaxNoVent", 0),
                            item.get("mmaxWithVent", 0),
                            no_val,
                            with_val,
                        ]
                        for idx, value in enumerate(values, start=1):
                            c = sheet.cell(row=r, column=idx, value=value)
                            c.border = border
                            if idx > 1:
                                c.alignment = center
                            if idx in {22, 23}:
                                if str(value).upper() == "YES":
                                    c.fill = yes_fill
                                elif str(value).upper() == "NO":
                                    c.fill = no_fill
                        r += 1

                    sheet.freeze_panes = f"A{header_row + 1}"
                    two_dec_cols = {2, 3, 4, 5, 7}
                    four_dec_cols = {8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21}
                    for rr in range(header_row + 1, r):
                        for cc in two_dec_cols:
                            sheet.cell(row=rr, column=cc).number_format = "0.00"
                        for cc in four_dec_cols:
                            sheet.cell(row=rr, column=cc).number_format = "0.0000"

                    for col_idx in range(1, end_col + 1):
                        letter = sheet.cell(row=header_row, column=col_idx).column_letter
                        max_len = max(
                            len(str(sheet.cell(row=rr, column=col_idx).value or ""))
                            for rr in range(1, max(r, header_row + 1))
                        )
                        sheet.column_dimensions[letter].width = min(30, max(11, max_len + 2))

                    lookup_sheet = workbook.create_sheet("Lookup Tables")
                    lookup_sheet.cell(row=1, column=1, value="Table 9-2: C by Refrigerant").font = Font(bold=True)
                    lookup_sheet.cell(row=2, column=1, value="Refrigerant").font = Font(bold=True)
                    lookup_sheet.cell(row=2, column=2, value="C").font = Font(bold=True)
                    lr = 3
                    for ref_name, c_val in [("R-32", 1.00), ("R-452B", 1.02), ("R-454A", 0.92), ("R-454B", 0.97), ("R-454C", 0.95), ("R-457A", 0.71)]:
                        lookup_sheet.cell(row=lr, column=1, value=ref_name)
                        lookup_sheet.cell(row=lr, column=2, value=c_val)
                        lr += 1
                    lr += 1
                    lookup_sheet.cell(row=lr, column=1, value="Table 9-3: M by Area").font = Font(bold=True)
                    lr += 1
                    lookup_sheet.cell(row=lr, column=1, value="Area (ft2)").font = Font(bold=True)
                    lookup_sheet.cell(row=lr, column=2, value="M (lb)").font = Font(bold=True)
                    lr += 1
                    table_93 = [
                        (100, 3.4), (125, 4.3), (150, 5.2), (175, 6.0), (200, 6.9), (225, 7.8), (250, 8.6), (275, 9.5), (300, 10.3),
                        (325, 11.2), (350, 12.1), (375, 12.9), (400, 13.8), (425, 14.6), (450, 15.5), (475, 16.4), (500, 17.2), (525, 18.1),
                        (550, 19.0), (575, 19.8), (600, 20.7), (625, 21.5), (650, 22.4), (675, 23.3), (700, 24.1), (725, 25.0), (750, 25.9),
                        (775, 26.7), (800, 27.6), (825, 28.4), (850, 29.3), (875, 30.2), (900, 31.0), (925, 31.9), (950, 32.7), (975, 33.6),
                        (1000, 34.5), (1025, 35.1),
                    ]
                    for area_val, m_val in table_93:
                        lookup_sheet.cell(row=lr, column=1, value=area_val)
                        lookup_sheet.cell(row=lr, column=2, value=m_val)
                        lr += 1
                    lr += 1
                    lookup_sheet.cell(row=lr, column=1, value="Table 9-4: Mv by Ventilation").font = Font(bold=True)
                    lr += 1
                    lookup_sheet.cell(row=lr, column=1, value="Ventilation (cfm)").font = Font(bold=True)
                    lookup_sheet.cell(row=lr, column=2, value="Mv (lb)").font = Font(bold=True)
                    lr += 1
                    table_94 = [
                        (20, 0.4), (40, 0.7), (60, 1.1), (80, 1.4), (100, 1.8), (120, 2.1), (140, 2.5), (160, 2.8), (180, 3.2), (200, 3.5),
                        (220, 4.2), (240, 4.6), (260, 5.0), (280, 5.4), (300, 5.8), (320, 6.2), (340, 6.5), (360, 6.9), (380, 7.3), (400, 7.3),
                    ]
                    for q_val, mv_val in table_94:
                        lookup_sheet.cell(row=lr, column=1, value=q_val)
                        lookup_sheet.cell(row=lr, column=2, value=mv_val)
                        lr += 1
                    lookup_sheet.column_dimensions["A"].width = 26
                    lookup_sheet.column_dimensions["B"].width = 16
                else:
                    sheet.title = "Refrigerant Compliance Report"
                    center = Alignment(horizontal="center", vertical="center")
                    title_fill = PatternFill(fill_type="solid", fgColor="D8E5F6")
                    r = 1
                    sheet.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                    t = sheet.cell(row=r, column=1, value="Refrigerant Compliance Report")
                    t.font = Font(bold=True, size=13)
                    t.alignment = center
                    t.fill = title_fill
                    r += 2

                    summary_rows = [
                        ("Project", project_info.get("projectName", "")),
                        ("Project Number", project_info.get("projectNumber", "")),
                        ("Client", project_info.get("clientName", "")),
                        ("Location", project_info.get("projectLocation", "")),
                        ("Revision", project_info.get("revision", "")),
                        ("Refrigerant Type", classification.get("refrigerantType", "")),
                        ("Safety Group", "A2L"),
                        ("System Type", classification.get("systemType", "")),
                        ("Space Type", classification.get("spaceType", "")),
                        ("Applicable Standard", result.get("applicableStandard", "")),
                        ("Data Version", classification.get("dataVersion", "")),
                        ("Method", result.get("methodLabel", "")),
                        ("Final Status", result.get("status", "")),
                        ("Exported At", now_text),
                    ]
                    for label, value in summary_rows:
                        sheet.cell(row=r, column=1, value=label).font = Font(bold=True)
                        sheet.cell(row=r, column=2, value=value)
                        r += 1
                    r += 1

                    sheet.cell(row=r, column=1, value="Input Summary").font = Font(bold=True, size=11)
                    r += 1
                    input_rows = [
                        ("Factory Charge (lb)", charge_inputs.get("factoryCharge", 0)),
                        ("Field Charge (lb)", charge_inputs.get("fieldCharge", 0)),
                        ("Installed Pipe Length (ft)", charge_inputs.get("installedPipeLength", 0)),
                        ("Included Pipe Length (ft)", charge_inputs.get("includedPipeLength", 0)),
                        ("Additional Charge Rate (lb/ft)", charge_inputs.get("additionalChargeRate", 0)),
                        ("Manual Additional Charge (lb)", charge_inputs.get("manualAdditionalCharge", 0) if charge_inputs.get("useManualAdditionalCharge") else 0),
                        ("Total System Charge (lb)", charge_result.get("ms", 0)),
                    ]
                    for label, value in input_rows:
                        sheet.cell(row=r, column=1, value=label)
                        sheet.cell(row=r, column=2, value=value)
                        r += 1
                    r += 1

                    applicable_standard = str(result.get("applicableStandard", "") or "")
                    if applicable_standard == "ASHRAE 15":
                        sheet.cell(row=r, column=1, value="ASHRAE 15 Inputs").font = Font(bold=True, size=11)
                        r += 1
                        ash15_rows = [
                            ("Equation Selection", ashrae15_inputs.get("manualEquation", ashrae15_inputs.get("selectionMode", ""))),
                            ("Connected Spaces", ashrae15_inputs.get("connectedSpaces", "")),
                            ("Air Circulation Provided", ashrae15_inputs.get("airCirculation", "")),
                            ("Room Area (ft2)", ashrae15_inputs.get("roomArea78", 0)),
                            ("Ceiling Height (ft)", ashrae15_inputs.get("ceilingHeight78", 0)),
                            ("LFL (lb/1000 ft3)", ashrae15_inputs.get("lfl78", 0)),
                            ("CF", ashrae15_inputs.get("cf78", 0)),
                            ("Focc", ashrae15_inputs.get("focc78", 0)),
                            ("Mdef (lb)", ashrae15_inputs.get("mdef79", 0)),
                            ("FLFL", ashrae15_inputs.get("flfl79", 0)),
                        ]
                        for label, value in ash15_rows:
                            sheet.cell(row=r, column=1, value=label)
                            sheet.cell(row=r, column=2, value=value)
                            r += 1
                    else:
                        sheet.cell(row=r, column=1, value="ASHRAE 15.2 Inputs").font = Font(bold=True, size=11)
                        r += 1
                        ash152_rows = [
                            ("Room Area (ft2)", ashrae152_inputs.get("roomArea", details.get("roomArea", 0))),
                            ("Dispersal/Space Height (ft)", ashrae152_inputs.get("spaceHeight", details.get("spaceHeight", 0))),
                            ("C", details.get("c", 0)),
                            ("M", details.get("m", 0)),
                            ("hc", details.get("hc", 0)),
                            ("Mc (lb)", details.get("mc", 0)),
                            ("Mv (lb)", details.get("mv", 0)),
                            ("m1 (lb)", details.get("m1", result.get("m1", 0))),
                            ("m2 (lb)", details.get("m2", result.get("m2", 0))),
                            ("mMAX No Vent (lb)", details.get("mmaxNoVent", result.get("limitValue", 0))),
                            ("mMAX With Vent (lb)", details.get("mmaxWithVent", 0)),
                        ]
                        if advanced_inputs.get("useMrel") == "YES":
                            ash152_rows.append(("mREL (lb)", advanced_inputs.get("mrel", details.get("mrel", 0))))
                        for label, value in ash152_rows:
                            sheet.cell(row=r, column=1, value=label)
                            sheet.cell(row=r, column=2, value=value)
                            r += 1
                    r += 1

                    sheet.cell(row=r, column=1, value="Result Summary").font = Font(bold=True, size=11)
                    r += 1
                    governing_charge = details.get("compareValue", charge_result.get("ms", 0))
                    governing_limit = result.get("limitValue", 0)
                    margin = (governing_limit or 0) - (governing_charge or 0)
                    conclusion = "Additional mitigation/manual review required."
                    if result.get("status") == "COMPLIANT":
                        conclusion = "System charge is within allowable limit."
                    elif result.get("status") == "NOT COMPLIANT":
                        conclusion = "System charge exceeds allowable limit."
                    elif result.get("status") == "INCOMPLETE":
                        conclusion = "Enter required inputs to evaluate compliance."
                    for label, value in [
                        ("Governing Charge (lb)", governing_charge),
                        ("Governing Limit (lb)", governing_limit),
                        ("Margin (lb)", margin),
                        ("Final Status", result.get("status", "")),
                        ("Conclusion", conclusion),
                    ]:
                        sheet.cell(row=r, column=1, value=label).font = Font(bold=(label in {"Final Status", "Conclusion"}))
                        sheet.cell(row=r, column=2, value=value)
                        r += 1

                    sheet.column_dimensions["A"].width = 40
                    sheet.column_dimensions["B"].width = 28
                data_rows = 1
            else:
                return {"ok": False, "reason": "Unsupported section export."}

            output_dir = Path.home() / "Downloads"
            output_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{section_label}_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = output_dir / filename
            workbook.save(output_path)
            return {
                "ok": True,
                "path": str(output_path),
                "rowCount": int(data_rows),
                "exportedAt": now_text,
            }
        except Exception as exc:
            return {"ok": False, "reason": f"Export failed: {exc}"}


def base_dir() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)  # type: ignore[attr-defined]
    return Path(__file__).resolve().parent


def app_url() -> str:
    index_path = base_dir() / "app" / "index.html"
    return index_path.resolve().as_uri()


def initial_window_geometry() -> dict:
    # Balanced default: comfortable first view without clipping.
    fallback = {"width": 1200, "height": 860, "x": None, "y": None}
    try:
        import tkinter as tk

        root = tk.Tk()
        root.withdraw()
        sw = int(root.winfo_screenwidth())
        sh = int(root.winfo_screenheight())
        root.destroy()

        width = max(1080, min(int(sw * 0.74), sw - 90))
        height = max(800, min(int(sh * 0.85), sh - 90))
        x = max(0, (sw - width) // 2)
        y = max(0, (sh - height) // 2)
        return {"width": width, "height": height, "x": x, "y": y}
    except Exception:
        return fallback


def main() -> None:
    api = Api()
    geometry = initial_window_geometry()
    window = webview.create_window(
        title="Mechanical and Plumbing Calculator",
        url=app_url(),
        js_api=api,
        width=geometry["width"],
        height=geometry["height"],
        x=geometry["x"],
        y=geometry["y"],
        min_size=(980, 760),
    )
    webview.start()


if __name__ == "__main__":
    main()

